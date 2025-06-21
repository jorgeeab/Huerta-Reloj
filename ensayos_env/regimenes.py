from openpyxl import load_workbook

class RegimenesManager:
    def __init__(self, archivo_regimenes):
        self.archivo_regimenes = archivo_regimenes

    def cargar_regimenes(self):
        workbook = load_workbook(self.archivo_regimenes)
        return workbook.sheetnames

    def agregar_regimen(self, regimen_name):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(regimen_name)
            sheet.append(["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades"])
            workbook.save(self.archivo_regimenes)

    def agregar_tarea(self, regimen, tarea_details):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            sheet.append(list(tarea_details.values()))
            workbook.save(self.archivo_regimenes)

    def modificar_tarea(self, regimen, fila, updated_values):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            for i, value in enumerate(updated_values.values(), start=1):
                sheet.cell(row=fila, column=i).value = value
            workbook.save(self.archivo_regimenes)

    def eliminar_tarea(self, regimen, fila):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            sheet.delete_rows(fila)
            workbook.save(self.archivo_regimenes)
