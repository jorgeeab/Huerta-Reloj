from openpyxl import load_workbook

class PlantasManager:
    def __init__(self, archivo_plantas):
        self.archivo_plantas = archivo_plantas

    def cargar_eras(self):
        workbook = load_workbook(self.archivo_plantas)
        return workbook.sheetnames

    def agregar_planta(self, planta_details, era):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            sheet.append([
                planta_details['Nombre de la Planta'],
                planta_details['Regimen'],
                planta_details['Dia Uno'],
                planta_details['Posición X'],
                planta_details['Posición Y'],
                planta_details['Posición Z'],
                planta_details['Velocidad de Agua']
            ])
            workbook.save(self.archivo_plantas)

    def modificar_planta(self, era, fila, updated_values):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            for i, value in enumerate(updated_values.values(), start=1):
                sheet.cell(row=fila, column=i).value = value
            workbook.save(self.archivo_plantas)

    def eliminar_planta(self, era, fila):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            sheet.delete_rows(fila)
            workbook.save(self.archivo_plantas)
