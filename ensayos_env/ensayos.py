from openpyxl import load_workbook, Workbook
import pandas as pd
from datetime import datetime, timedelta
import os

class EnsayosEnv:
    def __init__(self, archivo_plantas, archivo_regimenes, archivo_ensayos):
        self.archivo_plantas = archivo_plantas
        self.archivo_regimenes = archivo_regimenes
        self.archivo_ensayos = archivo_ensayos
        self._crear_archivo_si_no_existe(self.archivo_plantas)
        self._crear_archivo_si_no_existe(self.archivo_regimenes)
        self._crear_archivo_si_no_existe(self.archivo_ensayos)

    def _crear_archivo_si_no_existe(self, archivo):
        if not os.path.exists(archivo):
            wb = Workbook()
            wb.save(archivo)

    def crear_era(self, nombre_era):
        print(f"Intentando crear era {nombre_era}")
        workbook = load_workbook(self.archivo_plantas)
        if nombre_era not in workbook.sheetnames:
            sheet = workbook.create_sheet(nombre_era)
            sheet.append([
                'Nombre de la Planta',
                'Regimen',
                'Dia Uno',
                'Posición X',
                'Posición Y',
                'Posición Z',
                'Velocidad de Agua',
                'Detalles'  # Asegúrate de agregar la columna Detalles aquí
            ])
            workbook.save(self.archivo_plantas)
            print(f"Era {nombre_era} creada exitosamente.")
        else:
            print(f"La era {nombre_era} ya existe.")

    def agregar_planta(self, nombre_era, datos_planta):
        print(f"Intentando agregar planta a la era {nombre_era}")
        workbook = load_workbook(self.archivo_plantas)
        if nombre_era in workbook.sheetnames:
            sheet = workbook[nombre_era]
            print(f"Añadiendo la planta {datos_planta['Nombre de la Planta']} a la era {nombre_era}")
            sheet.append([
                datos_planta['Nombre de la Planta'],
                datos_planta['Regimen'],
                datos_planta['Dia Uno'],
                datos_planta['Posición X'],
                datos_planta['Posición Y'],
                datos_planta['Posición Z'],
                datos_planta['Velocidad de Agua'],
                datos_planta.get('Detalles', '')  # Incluye Detalles, con un valor predeterminado si no se proporciona
            ])
            workbook.save(self.archivo_plantas)
            print(f"Planta {datos_planta['Nombre de la Planta']} agregada correctamente.")
        else:
            raise ValueError(f"La era {nombre_era} no existe en el archivo {self.archivo_plantas}.")

    def modificar_planta(self, nombre_era, nombre_planta, datos_modificados):
        print(f"Intentando modificar la planta {nombre_planta} en la era {nombre_era}")
        workbook = load_workbook(self.archivo_plantas)
        if nombre_era in workbook.sheetnames:
            sheet = workbook[nombre_era]
            planta_encontrada = False
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value == nombre_planta:
                    planta_encontrada = True
                    print(f"Planta {nombre_planta} encontrada, actualizando datos.")
                    for key, value in datos_modificados.items():
                        idx = ['Nombre de la Planta', 'Regimen', 'Dia Uno', 'Posición X', 'Posición Y', 'Posición Z', 'Velocidad de Agua', 'Detalles'].index(key)
                        row[idx].value = value
                    workbook.save(self.archivo_plantas)
                    print(f"Planta {nombre_planta} modificada correctamente.")
                    return
            if not planta_encontrada:
                print(f"No se encontró la planta {nombre_planta} en la era {nombre_era}.")
                raise ValueError(f"La planta {nombre_planta} no se encontró en la era {nombre_era}.")
        else:
            raise ValueError(f"La era {nombre_era} no existe en el archivo {self.archivo_plantas}.")

    def eliminar_planta(self, nombre_era, nombre_planta):
        print(f"Intentando eliminar la planta {nombre_planta} de la era {nombre_era}")
        workbook = load_workbook(self.archivo_plantas)
        if nombre_era in workbook.sheetnames:
            sheet = workbook[nombre_era]
            planta_encontrada = False
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                if row[0].value == nombre_planta:
                    planta_encontrada = True
                    sheet.delete_rows(idx)
                    workbook.save(self.archivo_plantas)
                    print(f"Planta {nombre_planta} eliminada correctamente.")
                    return
            if not planta_encontrada:
                print(f"No se encontró la planta {nombre_planta} en la era {nombre_era}.")
                raise ValueError(f"La planta {nombre_planta} no se encontró en la era {nombre_era}.")
        else:
            raise ValueError(f"La era {nombre_era} no existe en el archivo {self.archivo_plantas}.")

    def agregar_regimen(self, nombre_regimen, tareas):
        print(f"Intentando agregar el régimen {nombre_regimen}")
        workbook = load_workbook(self.archivo_regimenes)
        if nombre_regimen not in workbook.sheetnames:
            sheet = workbook.create_sheet(nombre_regimen)
            sheet.append(["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades", "Detalles"])
            for tarea in tareas:
                sheet.append([
                    tarea["Tarea"],
                    tarea["Numero_Día"],
                    tarea["Hora"],
                    tarea["tiempo_ejecución(s)"],
                    tarea["magnitud"],
                    tarea["unidades"],
                    tarea.get("Detalles", "")  # Asegurarse de incluir Detalles aquí también
                ])
            workbook.save(self.archivo_regimenes)
            print(f"Régimen {nombre_regimen} agregado correctamente.")
        else:
            print(f"El régimen {nombre_regimen} ya existe.")
            raise ValueError(f"El régimen {nombre_regimen} ya existe en el archivo {self.archivo_regimenes}.")

    def modificar_tarea_regimen(self, nombre_regimen, indice_tarea, datos_modificados):
        print(f"Intentando modificar la tarea {indice_tarea} en el régimen {nombre_regimen}")
        workbook = load_workbook(self.archivo_regimenes)
        if nombre_regimen in workbook.sheetnames:
            sheet = workbook[nombre_regimen]
            row = sheet[indice_tarea + 2]  # +2 para saltar el encabezado
            for key, value in datos_modificados.items():
                idx = ["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades", "Detalles"].index(key)
                row[idx].value = value
            workbook.save(self.archivo_regimenes)
            print(f"Tarea {indice_tarea} en el régimen {nombre_regimen} modificada correctamente.")
        else:
            print(f"El régimen {nombre_regimen} no existe.")
            raise ValueError(f"El régimen {nombre_regimen} no existe en el archivo {self.archivo_regimenes}.")

    def eliminar_tarea_regimen(self, nombre_regimen, indice_tarea):
        print(f"Intentando eliminar la tarea {indice_tarea} en el régimen {nombre_regimen}")
        workbook = load_workbook(self.archivo_regimenes)
        if nombre_regimen in workbook.sheetnames:
            sheet = workbook[nombre_regimen]
            sheet.delete_rows(indice_tarea + 2)  # +2 para saltar el encabezado
            workbook.save(self.archivo_regimenes)
            print(f"Tarea {indice_tarea} en el régimen {nombre_regimen} eliminada correctamente.")
        else:
            print(f"El régimen {nombre_regimen} no existe.")
            raise ValueError(f"El régimen {nombre_regimen} no existe en el archivo {self.archivo_regimenes}.")

    def crear_ensayo(self):
        print(f"Creando un nuevo ensayo combinando plantas y regímenes.")
        plantas_df = pd.read_excel(self.archivo_plantas, sheet_name=None)  # Carga todas las hojas
        regimenes_df = pd.read_excel(self.archivo_regimenes, sheet_name=None)  # Carga todas las hojas

        plantas_total_df = pd.concat(plantas_df.values(), ignore_index=True)
        tareas_combinadas = []

        for regimen_name, df_regimen in regimenes_df.items():
            plantas_regimen = plantas_total_df[plantas_total_df['Regimen'] == regimen_name]

            for _, planta in plantas_regimen.iterrows():
                dia_inicio = datetime.strptime(planta['Dia Uno'], '%Y-%m-%d')
                for _, tarea in df_regimen.iterrows():
                    dia_tarea = dia_inicio + timedelta(days=tarea['Numero_Día'] - 1)
                    fecha_hora_tarea = datetime.strptime(f'{dia_tarea.date()} {tarea["Hora"]}', '%Y-%m-%d %H:%M')
                    tareas_combinadas.append({
                        'Nombre de la Planta': planta['Nombre de la Planta'],
                        'Día': dia_tarea,
                        'Hora': fecha_hora_tarea.strftime('%H:%M'),
                        'Tarea': tarea['Tarea'],
                        'Regimen': regimen_name,
                        'Magnitud': tarea['magnitud'],
                        'Unidades': tarea['unidades'],
                        'Detalles': tarea.get('Detalles', '')  # Incluye Detalles con un valor predeterminado
                    })

        tareas_df = pd.DataFrame(tareas_combinadas)
        tareas_df.sort_values(by=['Día', 'Hora'], inplace=True)

        with pd.ExcelWriter(self.archivo_ensayos, engine='openpyxl') as writer:
            tareas_df.to_excel(writer, index=False)
        print("Ensayo creado y guardado exitosamente.")
