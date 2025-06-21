import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import pandas as pd
from openpyxl import load_workbook

class Interfaz:

    def __init__(self, master, archivo_plantas, archivo_regimenes, archivo_ensayos):
            self.master = master
            self.master.title("Gestión de Ensayos")

            self.archivo_plantas = archivo_plantas
            self.archivo_regimenes = archivo_regimenes
            self.archivo_ensayos = archivo_ensayos

            self.eras = {}  # Inicializar eras como un diccionario vacío

            # Configurar la interfaz de usuario
            self.setup_ui()

    def setup_ui(self):
        # Crear panel de pestañas
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(expand=1, fill="both")

        # Añadir nueva pestaña para Ensayos
        self.ensayos_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.ensayos_frame, text="Ensayos")
        self.setup_ensayos_tab()

        # Crear la pestaña de Plantas
        self.plantas_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.plantas_frame, text="Plantas")
        self.setup_plantas_section()

        # Crear la pestaña de Regímenes
        self.regimenes_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.regimenes_frame, text="Regimenes")
        self.setup_regimenes_tab()

    def setup_ensayos_tab(self):
        self.ensayos_treeview = ttk.Treeview(self.ensayos_frame)
        self.ensayos_treeview.pack(side='left', expand=True, fill='both')

        self.ensayos_treeview['columns'] = ("Nombre de la Planta", "Día", "Hora", "Tarea", "Regimen", "Magnitud", "Unidades", "Detalles")
        self.ensayos_treeview['show'] = 'headings'

        for col in self.ensayos_treeview['columns']:
            self.ensayos_treeview.heading(col, text=col)
            self.ensayos_treeview.column(col, width=120)

        self.actualizar_treeview_ensayos()

    def setup_plantas_section(self):
        self.left_frame = ttk.Frame(self.plantas_frame, width=200)
        self.left_frame.pack_propagate(0)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.cargar_eras(self.left_frame)

        self.right_frame = ttk.Frame(self.plantas_frame)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame = ttk.Frame(self.right_frame)
        self.buttons_frame.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.buttons_frame, text="Agregar Planta", command=self.agregar_planta).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.buttons_frame, text="Modificar Planta", command=self.modificar_planta).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.buttons_frame, text="Eliminar Planta", command=self.eliminar_planta).pack(side=tk.LEFT, padx=5)

        self.plantas_treeview = ttk.Treeview(self.right_frame)
        self.plantas_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

    def setup_regimenes_tab(self):
        self.left_frame_regimenes = ttk.Frame(self.regimenes_frame, width=200)
        self.left_frame_regimenes.pack_propagate(0)
        self.left_frame_regimenes.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.right_frame_regimenes = ttk.Frame(self.regimenes_frame)
        self.right_frame_regimenes.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame_regimenes = ttk.Frame(self.right_frame_regimenes)
        self.buttons_frame_regimenes.pack(side=tk.TOP, fill=tk.X)
        self.agregar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Agregar Tarea",
                                               command=self.agregar_tarea_regimen)
        self.agregar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.modificar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Modificar Tarea", state='disabled',
                                                 command=self.modificar_tarea_regimen)
        self.modificar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.eliminar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Eliminar Tarea", state='disabled',
                                                command=self.eliminar_tarea_regimen)
        self.eliminar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.regimenes_treeview = ttk.Treeview(self.right_frame_regimenes)
        self.regimenes_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

        self.regimenes_treeview['columns'] = ("Tarea", "Numero_Día", "Hora", "Magnitud", "Unidades", "Detalles")
        self.regimenes_treeview['show'] = 'headings'

        for col in self.regimenes_treeview['columns']:
            self.regimenes_treeview.heading(col, text=col)
            self.regimenes_treeview.column(col, width=120)

        self.regimenes_treeview.bind("<<TreeviewSelect>>", self.on_regimen_treeview_select)

        self.cargar_regimenes()

    def actualizar_treeview_ensayos(self):
        tareas_df = pd.read_excel(self.archivo_ensayos)

        for i in self.ensayos_treeview.get_children():
            self.ensayos_treeview.delete(i)

        for _, fila in tareas_df.iterrows():
            self.ensayos_treeview.insert(
                '', 'end',
                values=(
                    fila['Nombre de la Planta'],
                    fila['Día'],
                    fila['Hora'],
                    fila['Tarea'],
                    fila['Regimen'],
                    fila['Magnitud'],
                    fila['Unidades'],
                    fila['Detalles']  # Asegurarse de incluir Detalles aquí
                )
            )

    def cargar_eras(self, container):
        for widget in container.winfo_children():
            widget.destroy()

        self.agregar_era_button = ttk.Button(container, text="Agregar Era", command=self.agregar_era, width=20)
        self.agregar_era_button.pack(side=tk.TOP, fill=tk.X)

        self.eliminar_era_button = ttk.Button(container, text="Eliminar Era", command=self.eliminar_era, width=20)
        self.eliminar_era_button.pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_plantas)

        for sheet_name in workbook.sheetnames:
            self.eras[sheet_name] = ttk.Button(container, text=sheet_name,
                                               command=lambda era=sheet_name: self.seleccionar_era(era))
            self.eras[sheet_name].pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def seleccionar_era(self, era):
        self.era_seleccionada = era
        self.mostrar_plantas_por_era(era)

    def mostrar_plantas_por_era(self, era):
        self.plantas_treeview.delete(*self.plantas_treeview.get_children())
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columns = ['Nombre de la Planta', 'Día Uno', 'Regimen', 'Velocidad de Agua', 'ÁnguloV', 'ÁnguloH', 'Posición X',
                       'Posición Y', 'Posición Z', 'Detalles']  # Asegúrate de incluir Detalles aquí
            self.plantas_treeview['columns'] = columns
            self.plantas_treeview['show'] = 'headings'
            for col in columns:
                self.plantas_treeview.heading(col, text=col)
                self.plantas_treeview.column(col, width=100)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.plantas_treeview.insert("", tk.END, values=row)
        else:
            messagebox.showerror("Error", f"La era '{era}' no tiene una hoja correspondiente en el libro de Excel.")

    def agregar_era(self):
        era_name = simpledialog.askstring("Nueva Era", "Nombre de la nueva Era:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name not in workbook.sheetnames:
                workbook.create_sheet(era_name)
                workbook.save(self.archivo_plantas)
                self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' ya existe.")

    def eliminar_era(self):
        era_name = simpledialog.askstring("Eliminar Era", "Nombre de la Era a eliminar:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Confirmar", f"¿Estás seguro de que deseas eliminar la era '{era_name}'?")
                if confirm:
                    del workbook[era_name]
                    workbook.save(self.archivo_plantas)
                    self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' no existe.")

    def agregar_planta(self):
        add_window = tk.Toplevel(self.master)
        add_window.title("Agregar Nueva Planta")
        valores_actuales = {
            'Nombre de la Planta': '',
            'Regimen': '',
            'Dia Uno': '2024-01-01',
            'Posición X': 0,
            'Posición Y': 0,
            'Posición Z': 0,
            'Velocidad de Agua': 0,
            'Detalles': ''  # Asegúrate de incluir Detalles aquí
        }
        entries = {}
        for idx, (field, value) in enumerate(valores_actuales.items()):
            label = ttk.Label(add_window, text=field)
            label.grid(row=idx, column=0, padx=10, pady=10)
            entry_var = tk.StringVar(add_window, value=value)
            entry = ttk.Entry(add_window, textvariable=entry_var)
            entry.grid(row=idx, column=1, padx=10, pady=10)
            entries[field] = entry_var

        ttk.Button(add_window, text="Guardar Planta",
                   command=lambda: self.guardar_nueva_planta(entries, add_window)).grid(row=len(valores_actuales),
                                                                                        column=0, columnspan=2)

    def guardar_nueva_planta(self, entries, window):
        planta_details = {field: entry.get() for field, entry in entries.items()}

        if all(planta_details.values()):
            self.agregar_planta_a_excel(planta_details, self.era_seleccionada)
            window.destroy()
            self.mostrar_plantas_por_era(self.era_seleccionada)
        else:
            messagebox.showwarning("Agregar Planta", "Todos los campos son obligatorios.")

    def agregar_planta_a_excel(self, planta_details, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            sheet.append([
                planta_details['Nombre de la Planta'],
                planta_details['Regimen'],
                planta_details['Dia Uno'],
                planta_details['Posición X'],
                planta_details['Posición Y'],
                planta_details['Posición Z'],
                planta_details['Velocidad de Agua'],
                planta_details['Detalles']  # Asegúrate de incluir Detalles aquí
            ])
            workbook.save(self.archivo_plantas)

    def modificar_planta(self):
        selected_item = self.plantas_treeview.selection()

        if not selected_item:
            messagebox.showwarning("Modificar Planta", "Por favor, selecciona una planta para modificar.")
            return

        planta_seleccionada = self.plantas_treeview.item(selected_item, 'values')

        edit_window = tk.Toplevel(self.master)
        edit_window.title("Modificar Planta")

        fields = ['Nombre de la Planta', 'Regimen', 'Dia Uno', 'Posición X', 'Posición Y', 'Posición Z',
                  'Velocidad de Agua', 'Detalles']  # Asegúrate de incluir Detalles aquí
        entries = {}

        for idx, field in enumerate(fields):
            label = ttk.Label(edit_window, text=field)
            label.grid(row=idx, column=0, padx=10, pady=10)
            entry = ttk.Entry(edit_window)
            entry.grid(row=idx, column=1, padx=10, pady=10)
            entry.insert(0, planta_seleccionada[idx])
            entries[field] = entry

        confirm_button = ttk.Button(edit_window, text="Confirmar Modificación",
                                    command=lambda: self.confirmar_modificacion_planta(selected_item, entries,
                                                                                       edit_window))
        confirm_button.grid(row=len(fields), column=0, columnspan=2)

    def confirmar_modificacion_planta(self, selected_item, entries, edit_window):
        updated_values = {field: entry.get() for field, entry in entries.items()}
        self.plantas_treeview.item(selected_item, values=list(updated_values.values()))
        self.actualizar_planta_en_excel(selected_item, updated_values)
        edit_window.destroy()

    def actualizar_planta_en_excel(self, selected_item, updated_values):
        workbook = load_workbook(self.archivo_plantas)
        sheet = workbook[self.era_seleccionada]
        fila_excel = self.plantas_treeview.index(selected_item) + 2
        col_order = ['Nombre de la Planta', 'Regimen', 'Dia Uno', 'Posición X', 'Posición Y', 'Posición Z',
                     'Velocidad de Agua', 'Detalles']  # Asegúrate de incluir Detalles aquí
        for i, field in enumerate(col_order, start=1):
            if field in updated_values:
                sheet.cell(row=fila_excel, column=i).value = updated_values[field]

        workbook.save(self.archivo_plantas)

    def eliminar_planta(self):
        selected_item = self.plantas_treeview.selection()
        if selected_item:
            respuesta = messagebox.askyesno("Eliminar Planta", "¿Estás seguro de que quieres eliminar esta planta?")
            if respuesta:
                self.eliminar_planta_de_excel(selected_item)
                self.mostrar_plantas_por_era(self.era_seleccionada)
        else:
            messagebox.showwarning("Eliminar Planta", "Por favor, selecciona una planta para eliminar.")

    def eliminar_planta_de_excel(self, selected_item):
        if self.era_seleccionada:
            workbook = load_workbook(self.archivo_plantas)
            if self.era_seleccionada in workbook.sheetnames:
                sheet = workbook[self.era_seleccionada]
                fila_para_borrar = self.plantas_treeview.index(selected_item) + 2
                sheet.delete_rows(fila_para_borrar)
                workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error", "No se ha seleccionado ninguna era.")

    def cargar_regimenes(self):
        for widget in self.left_frame_regimenes.winfo_children():
            widget.destroy()

        ttk.Button(self.left_frame_regimenes, text="Agregar Régimen",
                   command=self.agregar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.left_frame_regimenes, text="Eliminar Régimen",
                   command=self.eliminar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_regimenes)
        for regimen_name in workbook.sheetnames:
            ttk.Button(self.left_frame_regimenes, text=regimen_name,
                       command=lambda r=regimen_name: self.seleccionar_regimen(r),
                       width=20).pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def agregar_regimen(self):
        regimen_name = simpledialog.askstring("Nuevo Régimen", "Nombre del nuevo Régimen:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(regimen_name)
                sheet.append(["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades", "Detalles"])  # Asegúrate de incluir Detalles aquí
                workbook.save(self.archivo_regimenes)
                self.cargar_regimenes()
            else:
                messagebox.showerror("Error", f"El régimen '{regimen_name}' ya existe.")

    def eliminar_regimen(self):
        regimen_name = simpledialog.askstring("Eliminar Régimen", "Nombre del Régimen a eliminar:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Eliminar", f"¿Estás seguro de eliminar el régimen '{regimen_name}'?")
                if confirm:
                    del workbook[regimen_name]
                    workbook.save(self.archivo_regimenes)
                    self.cargar_regimenes()
            else:
                messagebox.showerror("Error", "Régimen no encontrado.")

    def seleccionar_regimen(self, regimen):
        self.regimen_actual = regimen
        self.mostrar_tareas_por_regimen(regimen)

    def agregar_tarea_regimen(self):
        regimen = self.regimen_actual

        ventana_emergente = tk.Toplevel(self.master)
        ventana_emergente.title("Agregar Nueva Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "Tiempo de ejecución (s)", "Magnitud", "Unidades", "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}

        for idx, campo in enumerate(campos):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, sticky="w")
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if regimen in workbook.sheetnames:
                sheet = workbook[regimen]
                sheet.append(list(detalles_tarea.values()))
                workbook.save(self.archivo_regimenes)
                ventana_emergente.destroy()
                self.mostrar_tareas_por_regimen(regimen)
            else:
                messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Agregar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                    columnspan=2, pady=10)

    def modificar_tarea_regimen(self):
        selected_item = self.regimenes_treeview.selection()[0]
        tarea_seleccionada = self.regimenes_treeview.item(selected_item, "values")

        ventana_emergente = tk.Toplevel(self.master)
        ventana_emergente.title("Modificar Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades", "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}
        for idx, (campo, valor) in enumerate(zip(campos, tarea_seleccionada)):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, padx=10, pady=5)
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entrada.insert(0, valor)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea_modificada = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if self.regimen_actual in workbook.sheetnames:
                sheet = workbook[self.regimen_actual]
                fila_tarea = self.regimenes_treeview.index(selected_item) + 2
                for idx, valor in enumerate(detalles_tarea_modificada.values(), start=1):
                    sheet.cell(row=fila_tarea, column=idx, value=valor)
                workbook.save(self.archivo_regimenes)
                self.regimenes_treeview.item(selected_item, values=list(detalles_tarea_modificada.values()))
                ventana_emergente.destroy()
            else:
                messagebox.showerror("Error", f"El régimen '{self.regimen_actual}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Modificar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                      columnspan=2, pady=10)

    def eliminar_tarea_regimen(self, regimen):
        selected = self.regimenes_treeview.selection()
        if selected:
            tarea_id = self.regimenes_treeview.index(selected[0]) + 2
            workbook = load_workbook(self.archivo_regimenes)
            sheet = workbook[regimen]
            sheet.delete_rows(tarea_id)
            workbook.save(self.archivo_regimenes)
            self.mostrar_tareas_por_regimen(regimen)

    def mostrar_tareas_por_regimen(self, regimen):
        try:
            if self.regimenes_treeview.winfo_exists():
                self.regimenes_treeview.delete(*self.regimenes_treeview.get_children())
                workbook = load_workbook(self.archivo_regimenes)
                if regimen in workbook.sheetnames:
                    sheet = workbook[regimen]
                    if not self.regimenes_treeview['columns']:
                        self.regimenes_treeview['columns'] = [sheet.cell(row=1, column=j).value for j in
                                                              range(1, sheet.max_column + 1)]
                        for col in self.regimenes_treeview['columns']:
                            self.regimenes_treeview.heading(col, text=col)
                            self.regimenes_treeview.column(col, width=100)
                    for i in range(2, sheet.max_row + 1):
                        row_values = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
                        self.regimenes_treeview.insert("", tk.END, values=row_values)
                else:
                    messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
            self.configurar_botones_accion_regimen(regimen)

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al mostrar las tareas: {e}")

    def configurar_botones_accion_regimen(self, regimen):
        if not hasattr(self, 'agregar_tarea_button'):
            self.agregar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Agregar Tarea a " + regimen,
                                                   command=lambda: self.agregar_tarea_regimen())
            self.agregar_tarea_button.pack(padx=10, pady=5)

            self.eliminar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Eliminar Tarea de " + regimen,
                                                    command=lambda: self.eliminar_tarea_regimen(regimen),
                                                    state='disabled')
            self.eliminar_tarea_button.pack(padx=10, pady=5)

            self.modificar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Modificar Tarea de " + regimen,
                                                     command=lambda: self.modificar_tarea_regimen(),
                                                     state='disabled')
            self.modificar_tarea_button.pack(padx=10, pady=5)
        else:
            self.agregar_tarea_button.config(text="Agregar Tarea a " + regimen,
                                             command=lambda: self.agregar_tarea_regimen())
            self.eliminar_tarea_button.config(text="Eliminar Tarea de " + regimen,
                                              command=lambda: self.eliminar_tarea_regimen(regimen))
            self.modificar_tarea_button.config(text="Modificar Tarea de " + regimen,
                                               command=lambda: self.modificar_tarea_regimen())

    def on_regimen_treeview_select(self, event):
        selected = self.regimenes_treeview.selection()
        if selected:
            self.eliminar_tarea_button['state'] = 'normal'
            self.modificar_tarea_button['state'] = 'normal'
        else:
            self.eliminar_tarea_button['state'] = 'disabled'
            self.modificar_tarea_button['state'] = 'disabled'

    def on_closing(self):
        self.master.destroy()
