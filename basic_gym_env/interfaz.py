import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import pandas as pd
from openpyxl import load_workbook
import threading
import time
from .basic_env import BasicEnv  # Asegúrate de que BasicEnv esté disponible para importación


class Interfaz(tk.Tk):
    def __init__(self, archivo_plantas, archivo_regimenes, archivo_ensayos, env=None):
        super().__init__()

        self.env = env  # Instancia del entorno BasicEnv

        self.title("Gestión de Ensayos y Control del Robot")
        self.archivo_plantas = archivo_plantas
        self.archivo_regimenes = archivo_regimenes
        self.archivo_ensayos = archivo_ensayos

        self.eras = {}
        self.entries = {}
        self.labels = {}
        self.buttons = {}

        self.setup_ui()
        self.bind_keys()  # Asocia teclas físicas del teclado

        # Hilo para recibir y procesar datos y ejecutar pasos en el ambiente
        self.update_thread = threading.Thread(target=self.update_data)
        self.update_thread.daemon = True
        self.update_thread.start()

    def setup_ui(self):
        # Crear panel de pestañas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=1, fill="both")

        # Añadir nueva pestaña para Ensayos
        self.ensayos_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.ensayos_frame, text="Ensayos")
        self.setup_ensayos_tab()

        # Crear la pestaña de Plantas
        self.plantas_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.plantas_frame, text="Plantas")
        self.setup_plantas_section()

        # Crear la pestaña de Regímenes
        self.regimenes_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.regimenes_frame, text="Regímenes")
        self.setup_regimenes_tab()

        # Paneles para control del robot
        self.robot_control_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.robot_control_frame, text="Control del Robot")
        self.setup_robot_control()

    def setup_robot_control(self):
        left_frame = ttk.Frame(self.robot_control_frame)
        left_frame.pack(side='left', fill='both', expand=True)

        right_frame = ttk.Frame(self.robot_control_frame)
        right_frame.pack(side='right', fill='both', expand=True)

        # Crear controles PID a la izquierda
        self.create_pid_controls(left_frame, 'PID Corredera')
        self.create_pid_controls(left_frame, 'PID Ángulo')
        self.create_pid_controls(left_frame, 'PID Válvula')

        # Crear controles para Setpoints (normal al inicio)
        self.create_control_entries(right_frame, 'Setpoint Corredera', 0, 400)
        self.create_control_entries(right_frame, 'Setpoint Ángulo', 0, 360)
        self.create_control_entries(right_frame, 'Setpoint Water', 0, 100)

        # Crear controles de Energía (desactivados al inicio)
        self.create_control_entries(right_frame, 'Energía Motor Corredera', -255, 255, initial_state='disabled')
        self.create_control_entries(right_frame, 'Energía Motor Ángulo', -255, 255, initial_state='disabled')
        self.create_control_entries(right_frame, 'Energía Motor Válvula', -255, 255, initial_state='disabled')

        # Crear botones de simulación de teclas
        self.create_keyboard_buttons(right_frame)

        # Modo manual
        self.manual_mode_var = tk.IntVar()
        self.manual_mode_var.trace_add('write', self.on_manual_mode_change)  # Añadir traza para detectar cambios
        self.manual_mode_check = tk.Checkbutton(left_frame, text="Modo Manual", variable=self.manual_mode_var,
                                                command=self.toggle_manual_mode)
        self.manual_mode_check.pack(pady=5)
        self.buttons['Modo Manual'] = self.manual_mode_check

        # Checkbox para activar/desactivar el joypad
        self.joypad_var = tk.IntVar(value=0)  # Inicia desactivado
        self.joypad_check = tk.Checkbutton(left_frame, text="Activar Joypad", variable=self.joypad_var,
                                           command=self.toggle_joypad)
        self.joypad_check.pack(pady=5)
        self.buttons['Activar Joypad'] = self.joypad_check

        # Cuadros de texto para mostrar datos enviados y recibidos
        self.text_font = ("TkFixedFont", 8)  # Configurar fuente con tamaño 8

        self.data_text_sent = tk.Text(self, height=5, width=100, font=self.text_font)
        self.data_text_sent.pack(pady=10)

        self.data_text_received = tk.Text(self, height=5, width=100, font=self.text_font)
        self.data_text_received.pack(pady=10)

    def toggle_joypad(self):
        """Habilita o deshabilita el joypad en el entorno según el estado del checkbox."""
        joypad_enabled = self.joypad_var.get()
        if self.env:
            if joypad_enabled:
                self.env.enable_joypad()
                self.data_text_sent.insert(tk.END, "Joypad habilitado.\n")
            else:
                self.env.disable_joypad()
                self.data_text_sent.insert(tk.END, "Joypad deshabilitado.\n")
            self.data_text_sent.see(tk.END)

    def create_control_entries(self, parent, label, min_value, max_value, initial_state='normal'):
        """Crea campos de entrada y botones de envío para setpoints y energías"""
        frame = tk.Frame(parent)
        frame.pack(fill='x', pady=5)

        # Etiqueta del nombre del control
        lbl = tk.Label(frame, text=label)
        lbl.pack(side='left')

        # Entrada para el valor del control
        entry = tk.Entry(frame, width=10, state=initial_state)  # El estado es controlado por 'initial_state'
        entry.pack(side='left', fill='x', expand=True)
        entry.insert(0, "0.00")  # Asegurarse que el valor tenga siempre dos decimales
        self.entries[label] = entry

        # Etiqueta para mostrar el valor actual
        self.labels[label] = tk.Label(frame, text="0.00")  # Siempre mostrar dos decimales
        self.labels[label].pack(side='left')

        # Botón de enviar para actualizar el valor
        btn_send = tk.Button(frame, text="Enviar", command=lambda l=label: self.update_entry_and_send_value(l))
        btn_send.pack(side='left')
        self.buttons[label] = btn_send

    def update_entry_and_send_value(self, label):
        """Actualiza el valor de la entrada asociada al label y lo envía al entorno"""
        try:
            value = float(self.entries[label].get())
            self.send_value(label, value)
        except ValueError:
            messagebox.showerror("Error", f"El valor de {label} no es válido.")

    def send_value(self, label, value):
        """Función para enviar el valor ajustado al entorno"""
        if self.env:
            # Llamamos a las funciones correctas del ambiente
            if label == 'Setpoint Corredera':
                self.env.set_corredera(value)
            elif label == 'Setpoint Ángulo':
                self.env.set_angulo(value)
            elif label == 'Setpoint Water':
                self.env.set_valvula(value)
            elif label == 'Energía Motor Corredera':
                self.env.set_energy_corredera(value)
            elif label == 'Energía Motor Ángulo':
                self.env.set_energy_angulo(value)
            elif label == 'Energía Motor Válvula':
                self.env.set_energy_valvula(value)

            # Ejecutar la acción en el entorno
            self.env.step()

            # Mostrar el comando enviado
            self.data_text_sent.insert(tk.END, f"Set {label} to {value:.2f}\n")
            self.data_text_sent.see(tk.END)

    def bind_keys(self):
        """Asocia las teclas físicas del teclado con las funciones."""
        self.bind('<Up>', lambda e: self.simulate_key_press('Up'))
        self.bind('<Down>', lambda e: self.simulate_key_press('Down'))
        self.bind('<Left>', lambda e: self.simulate_key_press('Left'))
        self.bind('<Right>', lambda e: self.simulate_key_press('Right'))
        self.bind('<w>', lambda e: self.simulate_key_press('w'))
        self.bind('<s>', lambda e: self.simulate_key_press('s'))

    def handle_key_press(self, key):
        """Controla qué sucede cuando se presiona una tecla"""
        if self.manual_mode_var.get():
            # Controlar las energías de los motores en modo manual
            if key == 'Up':
                self.adjust_value('Energía Motor Corredera', 1, -255, 255)
            elif key == 'Down':
                self.adjust_value('Energía Motor Corredera', -1, -255, 255)
            elif key == 'Right':
                self.adjust_value('Energía Motor Ángulo', 1, -255, 255)
            elif key == 'Left':
                self.adjust_value('Energía Motor Ángulo', -1, -255, 255)
            elif key == 'w':
                self.adjust_value('Energía Motor Válvula', 1, -255, 255)
            elif key == 's':
                self.adjust_value('Energía Motor Válvula', -1, -255, 255)
        else:
            # Controlar los setpoints en modo automático
            if key == 'Up':
                self.adjust_value('Setpoint Corredera', 10, 0, 400)
            elif key == 'Down':
                self.adjust_value('Setpoint Corredera', -10, 0, 400)
            elif key == 'Right':
                self.adjust_value('Setpoint Ángulo', 10, 0, 360)
            elif key == 'Left':
                self.adjust_value('Setpoint Ángulo', -10, 0, 360)
            elif key == 'w':
                self.adjust_value('Setpoint Water', 1, 0, 100)
            elif key == 's':
                self.adjust_value('Setpoint Water', -1, 0, 100)

    def simulate_key_release(self, key):
        """Simula visualmente la tecla soltada."""
        if key in self.key_buttons:
            self.key_buttons[key].config(relief="raised")  # Volver al estado original

    def create_keyboard_buttons(self, parent_frame):
        """Crea botones para simular teclas de flechas y W/S para controlar energías y setpoints."""
        button_frame = tk.Frame(parent_frame)
        button_frame.pack(pady=10)

        self.key_buttons = {}

        # Botones flechas arriba/abajo y izquierda/derecha
        self.create_key_button(button_frame, 'Up', '↑', row=0, column=1)
        self.create_key_button(button_frame, 'Down', '↓', row=2, column=1)
        self.create_key_button(button_frame, 'Left', '←', row=1, column=0)
        self.create_key_button(button_frame, 'Right', '→', row=1, column=2)

        # Botones W y S para control de fuerza
        self.create_key_button(button_frame, 'w', 'W', row=0, column=0)
        self.create_key_button(button_frame, 's', 'S', row=2, column=0)

    def simulate_key_press(self, key):
        """Simula la acción de presionar una tecla."""
        # Actualizar visualmente el botón
        if key in self.key_buttons:
            self.key_buttons[key].config(relief="sunken")  # Simula visualmente la tecla presionada

        # Manejar la acción asociada a la tecla
        self.handle_key_press(key)

        # Programar la simulación de la liberación de la tecla
        self.after(100, lambda: self.simulate_key_release(key))  # Simula la tecla soltada después de un retraso

    def create_key_button(self, parent_frame, key, text, row, column):
        """Crea un botón que simula una tecla del teclado."""
        frame = ttk.Frame(parent_frame)
        frame.grid(row=row, column=column, padx=5, pady=5)
        button = tk.Button(frame, text=text, width=5, height=2, command=lambda: self.simulate_key_press(key))
        button.pack()
        self.key_buttons[key] = button

    def adjust_value(self, label, delta, min_value, max_value):
        """Ajusta el valor actual de la entrada según el delta"""
        try:
            # Obtener el valor actual, si está vacío, asignar un valor por defecto de 0.0
            current_value = float(self.entries[label].get()) if self.entries[label].get() else 0.0
            # Ajustar el nuevo valor dentro de los límites
            new_value = max(min_value, min(current_value + delta, max_value))
            # Actualizar la entrada con el nuevo valor
            self.entries[label].delete(0, tk.END)
            self.entries[label].insert(0, f"{new_value:.2f}")  # Siempre con dos decimales
            # Enviar el valor actualizado
            self.send_value(label, new_value)
        except ValueError:
            # Si hay algún problema al convertir el valor, usar 0.0 como valor por defecto
            self.entries[label].delete(0, tk.END)
            self.entries[label].insert(0, "0.00")
            self.send_value(label, 0.0)

    def create_pid_controls(self, parent, label_prefix):
        """Crea controles de PID para Kp, Ki y Kd"""
        for pid_param in ['Kp', 'Ki', 'Kd']:
            full_label = f"{label_prefix} ({pid_param})"
            frame = tk.Frame(parent)
            frame.pack(fill='x', pady=5)

            lbl = tk.Label(frame, text=full_label)
            lbl.pack(side='left')

            entry = tk.Entry(frame, width=10)
            entry.pack(side='left', fill='x', expand=True)
            entry.insert(0, '0.0')
            self.entries[full_label] = entry

            btn = tk.Button(frame, text="Set", command=lambda l=full_label: self.send_pid_command([l]))
            btn.pack(side='left')

    def send_pid_command(self, label_set):
        """Envío de los parámetros PID al entorno"""
        try:
            # Obtener los valores actuales de los PIDs del entorno
            pid_corredera = list(self.env.current_action[3:6])  # PID para corredera
            pid_angulo = list(self.env.current_action[6:9])  # PID para ángulo
            pid_valvula = list(self.env.current_action[9:12])  # PID para válvula

            # Actualizar el PID correspondiente según el control que se está modificando
            for label in label_set:
                value = float(self.entries[label].get())
                if 'Corredera' in label:
                    if 'Kp' in label:
                        pid_corredera[0] = value
                    elif 'Ki' in label:
                        pid_corredera[1] = value
                    elif 'Kd' in label:
                        pid_corredera[2] = value
                elif 'Ángulo' in label:
                    if 'Kp' in label:
                        pid_angulo[0] = value
                    elif 'Ki' in label:
                        pid_angulo[1] = value
                    elif 'Kd' in label:
                        pid_angulo[2] = value
                elif 'Válvula' in label:
                    if 'Kp' in label:
                        pid_valvula[0] = value
                    elif 'Ki' in label:
                        pid_valvula[1] = value
                    elif 'Kd' in label:
                        pid_valvula[2] = value

            # Enviar los PIDs actualizados al entorno, manteniendo los valores anteriores
            if self.env:
                self.env.set_pid_corredera(*pid_corredera)
                self.env.set_pid_angulo(*pid_angulo)
                self.env.set_pid_valvula(*pid_valvula)

                # Ejecutar la acción en el entorno
                self.env.step()

            command_str = f"Setting PID values: Corredera {pid_corredera}, Ángulo {pid_angulo}, Válvula {pid_valvula}\n"
            self.data_text_sent.insert(tk.END, command_str)
            self.data_text_sent.see(tk.END)

        except ValueError as e:
            print(f"Invalid input: {e}")

    def toggle_manual_mode(self):
        """Se llama cuando el usuario interactúa con el Checkbutton de modo manual"""
        manual_mode = self.manual_mode_var.get()

        # Actualizar el modo manual en el entorno
        if self.env:
            self.env.set_manual_mode(manual_mode)

            # Ejecutar la acción en el entorno
            self.env.step()

        # Mostrar el comando enviado en la interfaz de texto
        self.data_text_sent.insert(tk.END,
                                   f"Modo Manual {'Activado' if manual_mode else 'Desactivado'} enviado al entorno.\n")
        self.data_text_sent.see(tk.END)

        # Actualizar la interfaz
        self.update_ui_for_manual_mode()

    def on_manual_mode_change(self, *args):
        """Se llama cada vez que cambia el valor de self.manual_mode_var"""
        # Actualizar la interfaz según el modo manual
        self.update_ui_for_manual_mode()

    def update_ui_for_manual_mode(self):
        """Actualiza los elementos de la interfaz según el estado del modo manual"""
        manual_mode = self.manual_mode_var.get()

        # Habilitar o deshabilitar botones y entradas según el modo manual
        if manual_mode:
            # Desactivar Setpoints y PID, activar energías
            for label in ['Setpoint Corredera', 'Setpoint Ángulo', 'Setpoint Water']:
                self.entries[label].config(state='disabled')
            for pid in ['PID Corredera', 'PID Ángulo', 'PID Válvula']:
                for param in ['Kp', 'Ki', 'Kd']:
                    self.entries[f'{pid} ({param})'].config(state='disabled')

            # Activar energía de los motores
            for label in ['Energía Motor Corredera', 'Energía Motor Ángulo', 'Energía Motor Válvula']:
                self.entries[label].config(state='normal')
                self.buttons[label].config(state='normal')  # Habilitar botones de envío para energía
        else:
            # Activar Setpoints y PID, desactivar energías
            for label in ['Setpoint Corredera', 'Setpoint Ángulo', 'Setpoint Water']:
                self.entries[label].config(state='normal')
            for pid in ['PID Corredera', 'PID Ángulo', 'PID Válvula']:
                for param in ['Kp', 'Ki', 'Kd']:
                    self.entries[f'{pid} ({param})'].config(state='normal')

            # Desactivar energía de los motores
            for label in ['Energía Motor Corredera', 'Energía Motor Ángulo', 'Energía Motor Válvula']:
                self.entries[label].config(state='disabled')
                self.buttons[label].config(state='disabled')  # Deshabilitar botones de envío para energía

    def update_data(self):
        """Hilo para recibir datos y actualizarlos en la interfaz"""
        while True:
            if self.env:
                # Ejecutar self.env.step en el hilo secundario
                obs, reward, done, _ = self.env.step()

                # Programar la actualización de la interfaz en el hilo principal
                self.after(0, self.process_observation, obs, reward)
            time.sleep(0.1)

    def process_observation(self, obs, reward):
        """Procesa la observación y actualiza la interfaz en el hilo principal"""
        # Actualizar la interfaz con las observaciones
        self.update_interface_with_obs(obs)

        # Actualizar el cuadro de texto con los datos recibidos
        self.data_text_received.insert(tk.END, f"Obs: {obs}, Reward: {reward}\n")
        self.data_text_received.see(tk.END)

    def update_labels(self, obs):
        """Actualiza las etiquetas con los valores recibidos del entorno"""
        try:
            # Mapear nombres de variables a índices
            variable_indices = {name: idx for idx, name in enumerate(self.env.variable_names)}

            # Diccionario que mapea los nombres de las variables a los nombres de los labels
            label_mappings = {
                'Setpoint Corredera': 'setpoint_corredera',
                'Setpoint Ángulo': 'setpoint_angle',
                'Setpoint Water': 'setpoint_water',
                'Energía Motor Corredera': 'energia_motor_corredera',
                'Energía Motor Ángulo': 'energia_motor_angulo',
                'Energía Motor Válvula': 'energia_motor_valvula'
            }

            for label_name, variable_name in label_mappings.items():
                idx = variable_indices.get(variable_name)
                if idx is not None and idx < len(obs):
                    value = obs[idx]
                    self.labels[label_name].config(text=f"{value:.2f}")
        except Exception as e:
            print(f"Error updating labels: {e}")

    # Aquí continuarían los métodos relacionados con Plantas, Regímenes y Ensayos...

    ## FUNCIONES RELACIONADAS CON ENSAYOS LUEGO

    def setup_ensayos_tab(self):
        self.ensayos_treeview = ttk.Treeview(self.ensayos_frame)
        self.ensayos_treeview.pack(side='left', expand=True, fill='both')

        self.ensayos_treeview['columns'] = (
            "Nombre de la Planta", "Día", "Hora", "Tarea", "Regimen", "Magnitud", "Unidades", "Detalles")
        self.ensayos_treeview['show'] = 'headings'

        for col in self.ensayos_treeview['columns']:
            self.ensayos_treeview.heading(col, text=col)
            self.ensayos_treeview.column(col, width=120)

        self.actualizar_treeview_ensayos()

    def setup_plantas_section(self):
        self.left_frame = ttk.Frame(self.plantas_frame, width=200)
        self.left_frame.pack_propagate(0)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.cargar_eras(self.left_frame)

        self.right_frame = ttk.Frame(self.plantas_frame)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame = ttk.Frame(self.right_frame)
        self.buttons_frame.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.buttons_frame, text="Agregar Planta", command=self.agregar_planta).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.buttons_frame, text="Modificar Planta", command=self.modificar_planta).pack(side=tk.LEFT,
                                                                                                    padx=5)
        ttk.Button(self.buttons_frame, text="Eliminar Planta", command=self.eliminar_planta).pack(side=tk.LEFT, padx=5)

        self.plantas_treeview = ttk.Treeview(self.right_frame)
        self.plantas_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

    def setup_regimenes_tab(self):
        self.left_frame_regimenes = ttk.Frame(self.regimenes_frame, width=200)
        self.left_frame_regimenes.pack_propagate(0)
        self.left_frame_regimenes.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.right_frame_regimenes = ttk.Frame(self.regimenes_frame)
        self.right_frame_regimenes.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame_regimenes = ttk.Frame(self.right_frame_regimenes)
        self.buttons_frame_regimenes.pack(side=tk.TOP, fill=tk.X)
        self.agregar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Agregar Tarea",
                                               command=self.agregar_tarea_regimen)
        self.agregar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.modificar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Modificar Tarea", state='disabled',
                                                 command=self.modificar_tarea_regimen)
        self.modificar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.eliminar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Eliminar Tarea", state='disabled',
                                                command=self.eliminar_tarea_regimen)
        self.eliminar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.regimenes_treeview = ttk.Treeview(self.right_frame_regimenes)
        self.regimenes_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

        self.regimenes_treeview['columns'] = ("Tarea", "Numero_Día", "Hora", "Magnitud", "Unidades", "Detalles")
        self.regimenes_treeview['show'] = 'headings'

        for col in self.regimenes_treeview['columns']:
            self.regimenes_treeview.heading(col, text=col)
            self.regimenes_treeview.column(col, width=120)

        self.regimenes_treeview.bind("<<TreeviewSelect>>", self.on_regimen_treeview_select)

        self.cargar_regimenes()


    def cargar_eras(self, container):
        for widget in container.winfo_children():
            widget.destroy()

        self.agregar_era_button = ttk.Button(container, text="Agregar Era", command=self.agregar_era, width=20)
        self.agregar_era_button.pack(side=tk.TOP, fill=tk.X)

        self.eliminar_era_button = ttk.Button(container, text="Eliminar Era", command=self.eliminar_era, width=20)
        self.eliminar_era_button.pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_plantas)

        for sheet_name in workbook.sheetnames:
            self.eras[sheet_name] = ttk.Button(container, text=sheet_name,
                                               command=lambda era=sheet_name: self.seleccionar_era(era))
            self.eras[sheet_name].pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def seleccionar_era(self, era):
        self.era_seleccionada = era
        self.mostrar_plantas_por_era(era)

    def mostrar_plantas_por_era(self, era):
        self.plantas_treeview.delete(*self.plantas_treeview.get_children())
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columns = ['Nombre de la Planta', 'Día Uno', 'Regimen', 'Velocidad de Agua', 'ÁnguloV', 'ÁnguloH',
                       'Posición X', 'Posición Y', 'Posición Z', 'Detalles']  # Asegúrate de incluir Detalles aquí
            self.plantas_treeview['columns'] = columns
            self.plantas_treeview['show'] = 'headings'
            for col in columns:
                self.plantas_treeview.heading(col, text=col)
                self.plantas_treeview.column(col, width=100)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.plantas_treeview.insert("", tk.END, values=row)
        else:
            messagebox.showerror("Error", f"La era '{era}' no tiene una hoja correspondiente en el libro de Excel.")

    def agregar_era(self):
        era_name = simpledialog.askstring("Nueva Era", "Nombre de la nueva Era:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name not in workbook.sheetnames:
                workbook.create_sheet(era_name)
                workbook.save(self.archivo_plantas)
                self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' ya existe.")

    def eliminar_era(self):
        era_name = simpledialog.askstring("Eliminar Era", "Nombre de la Era a eliminar:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Confirmar", f"¿Estás seguro de que deseas eliminar la era '{era_name}'?")
                if confirm:
                    del workbook[era_name]
                    workbook.save(self.archivo_plantas)
                    self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' no existe.")

    def agregar_planta(self):
        add_window = tk.Toplevel(self)
        add_window.title("Agregar Nueva Planta")
        valores_actuales = {
            'Nombre de la Planta': '',
            'Regimen': '',
            'Dia Uno': '2024-01-01',
            'Posición X': 0,
            'Posición Y': 0,
            'Posición Z': 0,
            'Velocidad de Agua': 0,
            'Detalles': ''  # Asegúrate de incluir Detalles aquí
        }
        entries = {}
        for idx, (field, value) in enumerate(valores_actuales.items()):
            label = ttk.Label(add_window, text=field)
            label.grid(row=idx, column=0, padx=10, pady=10)
            entry_var = tk.StringVar(add_window, value=value)
            entry = ttk.Entry(add_window, textvariable=entry_var)
            entry.grid(row=idx, column=1, padx=10, pady=10)
            entries[field] = entry_var

        ttk.Button(add_window, text="Guardar Planta",
                   command=lambda: self.guardar_nueva_planta(entries, add_window)).grid(row=len(valores_actuales),
                                                                                        column=0, columnspan=2)

    def guardar_nueva_planta(self, entries, window):
        planta_details = {field: entry.get() for field, entry in entries.items()}

        if all(planta_details.values()):
            self.agregar_planta_a_excel(planta_details, self.era_seleccionada)
            window.destroy()
            self.mostrar_plantas_por_era(self.era_seleccionada)
        else:
            messagebox.showwarning("Agregar Planta", "Todos los campos son obligatorios.")

    def agregar_planta_a_excel(self, planta_details, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            sheet.append([
                planta_details['Nombre de la Planta'],
                planta_details['Regimen'],
                planta_details['Dia Uno'],
                planta_details['Posición X'],
                planta_details['Posición Y'],
                planta_details['Posición Z'],
                planta_details['Velocidad de Agua'],
                planta_details['Detalles']  # Asegúrate de incluir Detalles aquí
            ])
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def modificar_planta(self):
        item = self.plantas_treeview.selection()
        if item:
            selected_plant = self.plantas_treeview.item(item, 'values')
            modificar_window = tk.Toplevel(self)
            modificar_window.title("Modificar Planta")
            entries = {}
            for idx, (field, value) in enumerate(zip(self.plantas_treeview['columns'], selected_plant)):
                label = ttk.Label(modificar_window, text=field)
                label.grid(row=idx, column=0, padx=10, pady=10)
                entry_var = tk.StringVar(modificar_window, value=value)
                entry = ttk.Entry(modificar_window, textvariable=entry_var)
                entry.grid(row=idx, column=1, padx=10, pady=10)
                entries[field] = entry_var

            ttk.Button(modificar_window, text="Guardar Cambios",
                       command=lambda: self.guardar_cambios_planta(entries, modificar_window, item)).grid(
                row=len(selected_plant), column=0, columnspan=2)

    def guardar_cambios_planta(self, entries, window, item):
        new_values = [entry.get() for entry in entries.values()]
        self.plantas_treeview.item(item, values=new_values)

        self.actualizar_planta_en_excel(new_values, self.era_seleccionada)
        window.destroy()

    def actualizar_planta_en_excel(self, new_values, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                if row[0].value == new_values[0]:  # Asumiendo que el nombre de la planta es el primer campo
                    for idx, cell in enumerate(row):
                        cell.value = new_values[idx]
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def eliminar_planta(self):
        item = self.plantas_treeview.selection()
        if item:
            planta = self.plantas_treeview.item(item, 'values')
            confirm = messagebox.askyesno("Eliminar Planta", f"¿Deseas eliminar la planta '{planta[0]}'?")
            if confirm:
                self.eliminar_planta_de_excel(planta[0], self.era_seleccionada)
                self.plantas_treeview.delete(item)

    def eliminar_planta_de_excel(self, planta_name, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                if row[0].value == planta_name:  # Asumiendo que el nombre de la planta es el primer campo
                    sheet.delete_rows(row[0].row, 1)
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def on_regimen_treeview_select(self, event):
        selected_item = self.regimenes_treeview.selection()
        if selected_item:
            self.modificar_tarea_button['state'] = 'normal'
            self.eliminar_tarea_button['state'] = 'normal'
        else:
            self.modificar_tarea_button['state'] = 'disabled'
            self.eliminar_tarea_button['state'] = 'disabled'

    def cargar_regimenes(self):
        for widget in self.left_frame_regimenes.winfo_children():
            widget.destroy()

        ttk.Button(self.left_frame_regimenes, text="Agregar Régimen",
                   command=self.agregar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.left_frame_regimenes, text="Eliminar Régimen",
                   command=self.eliminar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_regimenes)
        for regimen_name in workbook.sheetnames:
            ttk.Button(self.left_frame_regimenes, text=regimen_name,
                       command=lambda r=regimen_name: self.seleccionar_regimen(r),
                       width=20).pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def agregar_regimen(self):
        regimen_name = simpledialog.askstring("Nuevo Régimen", "Nombre del nuevo Régimen:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(regimen_name)
                sheet.append(["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades",
                              "Detalles"])  # Asegúrate de incluir Detalles aquí
                workbook.save(self.archivo_regimenes)
                self.cargar_regimenes()
            else:
                messagebox.showerror("Error", f"El régimen '{regimen_name}' ya existe.")

    def eliminar_regimen(self):
        regimen_name = simpledialog.askstring("Eliminar Régimen", "Nombre del Régimen a eliminar:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Eliminar", f"¿Estás seguro de eliminar el régimen '{regimen_name}'?")
                if confirm:
                    del workbook[regimen_name]
                    workbook.save(self.archivo_regimenes)
                    self.cargar_regimenes()
            else:
                messagebox.showerror("Error", "Régimen no encontrado.")

    def seleccionar_regimen(self, regimen):
        self.regimen_actual = regimen
        self.mostrar_tareas_por_regimen(regimen)

    def agregar_tarea_regimen(self):
        regimen = self.regimen_actual

        ventana_emergente = tk.Toplevel(self)
        ventana_emergente.title("Agregar Nueva Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "Tiempo de ejecución (s)", "Magnitud", "Unidades",
                  "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}

        for idx, campo in enumerate(campos):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, sticky="w")
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if regimen in workbook.sheetnames:
                sheet = workbook[regimen]
                sheet.append(list(detalles_tarea.values()))
                workbook.save(self.archivo_regimenes)
                ventana_emergente.destroy()
                self.mostrar_tareas_por_regimen(regimen)
            else:
                messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Agregar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                    columnspan=2, pady=10)

    def modificar_tarea_regimen(self):
        selected_item = self.regimenes_treeview.selection()[0]
        tarea_seleccionada = self.regimenes_treeview.item(selected_item, "values")

        ventana_emergente = tk.Toplevel(self)
        ventana_emergente.title("Modificar Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades",
                  "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}
        for idx, (campo, valor) in enumerate(zip(campos, tarea_seleccionada)):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, padx=10, pady=5)
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entrada.insert(0, valor)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea_modificada = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if self.regimen_actual in workbook.sheetnames:
                sheet = workbook[self.regimen_actual]
                fila_tarea = self.regimenes_treeview.index(selected_item) + 2
                for idx, valor in enumerate(detalles_tarea_modificada.values(), start=1):
                    sheet.cell(row=fila_tarea, column=idx, value=valor)
                workbook.save(self.archivo_regimenes)
                self.regimenes_treeview.item(selected_item, values=list(detalles_tarea_modificada.values()))
                ventana_emergente.destroy()
            else:
                messagebox.showerror("Error", f"El régimen '{self.regimen_actual}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Modificar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                      columnspan=2, pady=10)

    def eliminar_tarea_regimen(self, regimen):
        selected = self.regimenes_treeview.selection()
        if selected:
            tarea_id = self.regimenes_treeview.index(selected[0]) + 2
            workbook = load_workbook(self.archivo_regimenes)
            sheet = workbook[regimen]
            sheet.delete_rows(tarea_id)
            workbook.save(self.archivo_regimenes)
            self.mostrar_tareas_por_regimen(regimen)

    def mostrar_tareas_por_regimen(self, regimen):
        try:
            if self.regimenes_treeview.winfo_exists():
                self.regimenes_treeview.delete(*self.regimenes_treeview.get_children())
                workbook = load_workbook(self.archivo_regimenes)
                if regimen in workbook.sheetnames:
                    sheet = workbook[regimen]
                    if not self.regimenes_treeview['columns']:
                        self.regimenes_treeview['columns'] = [sheet.cell(row=1, column=j).value for j in
                                                              range(1, sheet.max_column + 1)]
                        for col in self.regimenes_treeview['columns']:
                            self.regimenes_treeview.heading(col, text=col)
                            self.regimenes_treeview.column(col, width=100)
                    for i in range(2, sheet.max_row + 1):
                        row_values = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
                        self.regimenes_treeview.insert("", tk.END, values=row_values)
                else:
                    messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
            self.configurar_botones_accion_regimen(regimen)

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al mostrar las tareas: {e}")

    def configurar_botones_accion_regimen(self, regimen):
        if not hasattr(self, 'agregar_tarea_button'):
            self.agregar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Agregar Tarea a " + regimen,
                                                   command=lambda: self.agregar_tarea_regimen())
            self.agregar_tarea_button.pack(padx=10, pady=5)

            self.eliminar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Eliminar Tarea de " + regimen,
                                                    command=lambda: self.eliminar_tarea_regimen(regimen),
                                                    state='disabled')
            self.eliminar_tarea_button.pack(padx=10, pady=5)

            self.modificar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Modificar Tarea de " + regimen,
                                                     command=lambda: self.modificar_tarea_regimen(),
                                                     state='disabled')
            self.modificar_tarea_button.pack(padx=10, pady=5)
        else:
            self.agregar_tarea_button.config(text="Agregar Tarea a " + regimen,
                                             command=lambda: self.agregar_tarea_regimen())
            self.eliminar_tarea_button.config(text="Eliminar Tarea de " + regimen,
                                              command=lambda: self.eliminar_tarea_regimen(regimen))
            self.modificar_tarea_button.config(text="Modificar Tarea de " + regimen,
                                               command=lambda: self.modificar_tarea_regimen())

    def on_regimen_treeview_select(self, event):
        selected = self.regimenes_treeview.selection()
        if selected:
            self.eliminar_tarea_button['state'] = 'normal'
            self.modificar_tarea_button['state'] = 'normal'
        else:
            self.eliminar_tarea_button['state'] = 'disabled'
            self.modificar_tarea_button['state'] = 'disabled'

    def actualizar_treeview_ensayos(self):
        tareas_df = pd.read_excel(self.archivo_ensayos)

        for i in self.ensayos_treeview.get_children():
            self.ensayos_treeview.delete(i)

        for _, fila in tareas_df.iterrows():
            self.ensayos_treeview.insert(
                '', 'end',
                values=(
                    fila['Nombre de la Planta'],
                    fila['Día'],
                    fila['Hora'],
                    fila['Tarea'],
                    fila['Regimen'],
                    fila['Magnitud'],
                    fila['Unidades'],
                    fila['Detalles']  # Asegurarse de incluir Detalles aquí
                )
            )

if __name__ == "__main__":
    env = BasicEnv()  # Crea la instancia del entorno BasicEnv
    app = Interfaz('archivo_plantas.xlsx', 'archivo_regimenes.xlsx', 'archivo_ensayos.xlsx', env=env)
    app.mainloop()
