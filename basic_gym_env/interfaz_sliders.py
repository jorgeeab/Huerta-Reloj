import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import pandas as pd
from openpyxl import load_workbook
import threading
import time
from .basic_env import BasicEnv  # Asegúrate de que BasicEnv esté disponible para importación


class Interfaz(tk.Tk):
    def __init__(self, archivo_plantas, archivo_regimenes, archivo_ensayos, env=None):
        super().__init__()

        self.env = env  # Instancia del entorno BasicEnv
        self.lock = threading.Lock()  # Crear un bloqueo para la sincronización

        self.title("Gestión de Ensayos y Control del Robot")
        self.archivo_plantas = archivo_plantas
        self.archivo_regimenes = archivo_regimenes
        self.archivo_ensayos = archivo_ensayos

        self.eras = {}
        self.entries = {}
        self.labels = {}
        self.buttons = {}
        self.sliders = {}

        self.setup_ui()

        # Hilo para recibir y procesar datos y ejecutar pasos en el ambiente
        self.update_thread = threading.Thread(target=self.update_data)
        self.update_thread.daemon = True
        self.update_thread.start()

    def setup_ui(self):
        # Crear panel de pestañas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=1, fill="both")

        # Añadir nueva pestaña para Ensayos
        self.ensayos_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.ensayos_frame, text="Ensayos")
        self.setup_ensayos_tab()

        # Crear la pestaña de Plantas
        self.plantas_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.plantas_frame, text="Plantas")
        self.setup_plantas_section()

        # Crear la pestaña de Regímenes
        self.regimenes_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.regimenes_frame, text="Regimenes")
        self.setup_regimenes_tab()

        # Paneles para control del robot
        self.robot_control_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.robot_control_frame, text="Control del Robot")
        self.setup_robot_control()

    def update_data(self):
        while True:
            if self.env:
                # Ejecutar un paso en el ambiente y obtener las observaciones
                obs, reward, done, _ = self.env.step(self.env.current_action)
                self.lock.acquire()
                try:
                    # Actualizar la interfaz con las observaciones
                    self.update_interface_with_obs(obs)
                    # Mostrar datos en el cuadro de texto de recibidos
                    self.data_text_received.insert(tk.END, f"Obs: {obs}, Reward: {reward}\n")
                    self.data_text_received.see(tk.END)
                finally:
                    self.lock.release()
            time.sleep(0.1)  # Ajusta el tiempo según sea necesario

    def update_interface_with_obs(self, obs):
        try:
            # Imprimir los datos de obs para depuración
            print(f"Observations received: {obs}")
            # Asegurarse de que todas las observaciones sean floats antes de formatear
            obs = [float(x) if isinstance(x, (int, float)) else 0.0 for x in obs]
            # Actualizar todas las etiquetas con los valores correspondientes de las observaciones
            self.labels['Setpoint Corredera'].config(
                text=f"{obs[3]:.2f}" if isinstance(obs[3], (int, float)) else "N/A")
            self.labels['Setpoint Ángulo'].config(text=f"{obs[4]:.2f}" if isinstance(obs[4], (int, float)) else "N/A")
            self.labels['Setpoint Water'].config(text=f"{obs[5]:.2f}" if isinstance(obs[5], (int, float)) else "N/A")
            self.labels['PID Kp (Corredera)'].config(
                text=f"{obs[6]:.2f}" if isinstance(obs[6], (int, float)) else "N/A")
            self.labels['PID Ki (Corredera)'].config(
                text=f"{obs[7]:.2f}" if isinstance(obs[7], (int, float)) else "N/A")
            self.labels['PID Kd (Corredera)'].config(
                text=f"{obs[8]:.2f}" if isinstance(obs[8], (int, float)) else "N/A")
            self.labels['PID Kp (Ángulo)'].config(text=f"{obs[9]:.2f}" if isinstance(obs[9], (int, float)) else "N/A")
            self.labels['PID Ki (Ángulo)'].config(text=f"{obs[10]:.2f}" if isinstance(obs[10], (int, float)) else "N/A")
            self.labels['PID Kd (Ángulo)'].config(text=f"{obs[11]:.2f}" if isinstance(obs[11], (int, float)) else "N/A")
            self.labels['PID Kp (Válvula)'].config(
                text=f"{obs[12]:.2f}" if isinstance(obs[12], (int, float)) else "N/A")
            self.labels['PID Ki (Válvula)'].config(
                text=f"{obs[13]:.2f}" if isinstance(obs[13], (int, float)) else "N/A")
            self.labels['PID Kd (Válvula)'].config(
                text=f"{obs[14]:.2f}" if isinstance(obs[14], (int, float)) else "N/A")
            self.labels['Energía Motor Corredera'].config(
                text=f"{obs[16]:.2f}" if isinstance(obs[16], (int, float)) else "N/A")
            self.labels['Energía Motor Ángulo'].config(
                text=f"{obs[17]:.2f}" if isinstance(obs[17], (int, float)) else "N/A")
            self.labels['Energía Motor Válvula'].config(
                text=f"{obs[18]:.2f}" if isinstance(obs[18], (int, float)) else "N/A")
        except Exception as e:
            print(f"Error updating interface with obs: {e}")

    def setup_robot_control(self):

        left_frame = ttk.Frame(self.robot_control_frame)
        left_frame.pack(side='left', fill='both', expand=True)

        right_frame = ttk.Frame(self.robot_control_frame)
        right_frame.pack(side='right', fill='y', expand=True)

        self.sliders = {}
        self.labels = {}

        # Slider para Setpoint Corredera
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl = tk.Label(frame, text='Setpoint Corredera')
        lbl.pack(side='left')
        slider = tk.Scale(frame, from_=0, to=400, orient='horizontal',
                          command=lambda value: self.handle_slider_change('Setpoint Corredera', value))
        slider.pack(side='left', fill='x', expand=True)
        self.sliders['Setpoint Corredera'] = slider
        self.labels['Setpoint Corredera'] = tk.Label(frame, text="0")
        self.labels['Setpoint Corredera'].pack(side='left')

        # Slider para Setpoint Ángulo
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl = tk.Label(frame, text='Setpoint Ángulo')
        lbl.pack(side='left')
        slider = tk.Scale(frame, from_=0, to=360, orient='horizontal',
                          command=lambda value: self.handle_slider_change('Setpoint Ángulo', value))
        slider.pack(side='left', fill='x', expand=True)
        self.sliders['Setpoint Ángulo'] = slider
        self.labels['Setpoint Ángulo'] = tk.Label(frame, text="0")
        self.labels['Setpoint Ángulo'].pack(side='left')

        # Slider para Setpoint Water
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl = tk.Label(frame, text='Setpoint Water')
        lbl.pack(side='left')
        slider = tk.Scale(frame, from_=0, to=100, orient='horizontal',
                          command=lambda value: self.handle_slider_change('Setpoint Water', value))
        slider.pack(side='left', fill='x', expand=True)
        self.sliders['Setpoint Water'] = slider
        self.labels['Setpoint Water'] = tk.Label(frame, text="0")
        self.labels['Setpoint Water'].pack(side='left')

        # Slider para Energía Motor Corredera
        frame = tk.Frame(right_frame)
        frame.pack(fill='y', pady=5)
        lbl = tk.Label(frame, text='Energía Motor Corredera')
        lbl.pack()
        slider = tk.Scale(frame, from_=-255, to=255, orient='vertical',
                          command=lambda value: self.handle_slider_change('Energía Motor Corredera', value))
        slider.pack(fill='y', expand=True)
        self.sliders['Energía Motor Corredera'] = slider
        self.labels['Energía Motor Corredera'] = tk.Label(frame, text="0")
        self.labels['Energía Motor Corredera'].pack()

        # Slider para Energía Motor Ángulo
        frame = tk.Frame(right_frame)
        frame.pack(fill='y', pady=5)
        lbl = tk.Label(frame, text='Energía Motor Ángulo')
        lbl.pack()
        slider = tk.Scale(frame, from_=-255, to=255, orient='vertical',
                          command=lambda value: self.handle_slider_change('Energía Motor Ángulo', value))
        slider.pack(fill='y', expand=True)
        self.sliders['Energía Motor Ángulo'] = slider
        self.labels['Energía Motor Ángulo'] = tk.Label(frame, text="0")
        self.labels['Energía Motor Ángulo'].pack()

        # Slider para Energía Motor Válvula
        frame = tk.Frame(right_frame)
        frame.pack(fill='y', pady=5)
        lbl = tk.Label(frame, text='Energía Motor Válvula')
        lbl.pack()
        slider = tk.Scale(frame, from_=-255, to=255, orient='vertical',
                          command=lambda value: self.handle_slider_change('Energía Motor Válvula', value))
        slider.pack(fill='y', expand=True)
        self.sliders['Energía Motor Válvula'] = slider
        self.labels['Energía Motor Válvula'] = tk.Label(frame, text="0")
        self.labels['Energía Motor Válvula'].pack()

        # Controles adicionales (PID, modo manual, calibración)
        # PID Corredera
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl_kp = tk.Label(frame, text='PID Kp (Corredera)')
        lbl_kp.pack(side='left')
        entry_kp = tk.Entry(frame, width=10)
        entry_kp.pack(side='left', fill='x', expand=True)
        entry_kp.insert(0, '0.0')
        self.entries['PID Kp (Corredera)'] = entry_kp
        self.labels['PID Kp (Corredera)'] = tk.Label(frame, text="0")
        self.labels['PID Kp (Corredera)'].pack(side='left')

        lbl_ki = tk.Label(frame, text='PID Ki (Corredera)')
        lbl_ki.pack(side='left')
        entry_ki = tk.Entry(frame, width=10)
        entry_ki.pack(side='left', fill='x', expand=True)
        entry_ki.insert(0, '0.0')
        self.entries['PID Ki (Corredera)'] = entry_ki
        self.labels['PID Ki (Corredera)'] = tk.Label(frame, text="0")
        self.labels['PID Ki (Corredera)'].pack(side='left')

        lbl_kd = tk.Label(frame, text='PID Kd (Corredera)')
        lbl_kd.pack(side='left')
        entry_kd = tk.Entry(frame, width=10)
        entry_kd.pack(side='left', fill='x', expand=True)
        entry_kd.insert(0, '0.0')
        self.entries['PID Kd (Corredera)'] = entry_kd
        self.labels['PID Kd (Corredera)'] = tk.Label(frame, text="0")
        self.labels['PID Kd (Corredera)'].pack(side='left')

        btn = tk.Button(frame, text="Set", command=lambda: self.send_pid_command(
            ['PID Kp (Corredera)', 'PID Ki (Corredera)', 'PID Kd (Corredera)']))
        btn.pack(side='left')
        self.buttons['PID Corredera'] = btn

        # PID Ángulo
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl_kp = tk.Label(frame, text='PID Kp (Ángulo)')
        lbl_kp.pack(side='left')
        entry_kp = tk.Entry(frame, width=10)
        entry_kp.pack(side='left', fill='x', expand=True)
        entry_kp.insert(0, '0.0')
        self.entries['PID Kp (Ángulo)'] = entry_kp
        self.labels['PID Kp (Ángulo)'] = tk.Label(frame, text="0")
        self.labels['PID Kp (Ángulo)'].pack(side='left')

        lbl_ki = tk.Label(frame, text='PID Ki (Ángulo)')
        lbl_ki.pack(side='left')
        entry_ki = tk.Entry(frame, width=10)
        entry_ki.pack(side='left', fill='x', expand=True)
        entry_ki.insert(0, '0.0')
        self.entries['PID Ki (Ángulo)'] = entry_ki
        self.labels['PID Ki (Ángulo)'] = tk.Label(frame, text="0")
        self.labels['PID Ki (Ángulo)'].pack(side='left')

        lbl_kd = tk.Label(frame, text='PID Kd (Ángulo)')
        lbl_kd.pack(side='left')
        entry_kd = tk.Entry(frame, width=10)
        entry_kd.pack(side='left', fill='x', expand=True)
        entry_kd.insert(0, '0.0')
        self.entries['PID Kd (Ángulo)'] = entry_kd
        self.labels['PID Kd (Ángulo)'] = tk.Label(frame, text="0")
        self.labels['PID Kd (Ángulo)'].pack(side='left')

        btn = tk.Button(frame, text="Set", command=lambda: self.send_pid_command(
            ['PID Kp (Ángulo)', 'PID Ki (Ángulo)', 'PID Kd (Ángulo)']))
        btn.pack(side='left')
        self.buttons['PID Ángulo'] = btn

        # PID Válvula
        frame = tk.Frame(left_frame)
        frame.pack(fill='x', pady=5)
        lbl_kp = tk.Label(frame, text='PID Kp (Válvula)')
        lbl_kp.pack(side='left')
        entry_kp = tk.Entry(frame, width=10)
        entry_kp.pack(side='left', fill='x', expand=True)
        entry_kp.insert(0, '0.0')
        self.entries['PID Kp (Válvula)'] = entry_kp
        self.labels['PID Kp (Válvula)'] = tk.Label(frame, text="0")
        self.labels['PID Kp (Válvula)'].pack(side='left')

        lbl_ki = tk.Label(frame, text='PID Ki (Válvula)')
        lbl_ki.pack(side='left')
        entry_ki = tk.Entry(frame, width=10)
        entry_ki.pack(side='left', fill='x', expand=True)
        entry_ki.insert(0, '0.0')
        self.entries['PID Ki (Válvula)'] = entry_ki
        self.labels['PID Ki (Válvula)'] = tk.Label(frame, text="0")
        self.labels['PID Ki (Válvula)'].pack(side='left')

        lbl_kd = tk.Label(frame, text='PID Kd (Válvula)')
        lbl_kd.pack(side='left')
        entry_kd = tk.Entry(frame, width=10)
        entry_kd.pack(side='left', fill='x', expand=True)
        entry_kd.insert(0, '0.0')
        self.entries['PID Kd (Válvula)'] = entry_kd
        self.labels['PID Kd (Válvula)'] = tk.Label(frame, text="0")
        self.labels['PID Kd (Válvula)'].pack(side='left')

        btn = tk.Button(frame, text="Set", command=lambda: self.send_pid_command(
            ['PID Kp (Válvula)', 'PID Ki (Válvula)', 'PID Kd (Válvula)']))
        btn.pack(side='left')
        self.buttons['PID Válvula'] = btn

        self.manual_mode_var = tk.IntVar()
        self.manual_mode_check = tk.Checkbutton(left_frame, text="Modo Manual", variable=self.manual_mode_var,
                                                command=self.toggle_manual_mode)
        self.manual_mode_check.pack(pady=5)
        self.buttons['Modo Manual'] = self.manual_mode_check

        self.calibrate_button = tk.Button(left_frame, text="Calibrar Brújula", command=self.calibrate_compass)
        self.calibrate_button.pack(pady=5)
        self.buttons['Calibrar Brújula'] = self.calibrate_button

        self.reset_button = tk.Button(left_frame, text="Reset", command=self.reset_robot)
        self.reset_button.pack(pady=5)
        self.buttons['Reset'] = self.reset_button

        # Cuadros de texto para mostrar datos enviados y recibidos
        self.text_font = ("TkFixedFont", 8)  # Configurar fuente con tamaño 8

        self.data_text_sent = tk.Text(self, height=5, width=100, font=self.text_font)
        self.data_text_sent.pack(pady=10)

        self.data_text_received = tk.Text(self, height=5, width=100, font=self.text_font)
        self.data_text_received.pack(pady=10)

    def handle_slider_change(self, label, value, update_env=True):
        try:
            # Verifica el valor recibido
            print(f"Slider {label} moved to value: {value}")

            value = float(value)  # Asegúrate de que se puede convertir a float
            print(f"Converted value: {value}")

            # Actualizar la etiqueta correspondiente con el valor del slider
            self.labels[label].config(text=f"{value:.2f}")

            if update_env and self.env:
                action = self.env.current_action.copy()  # Copiar la acción actual del entorno

                # Actualizar la acción del entorno según el slider que se ha movido
                if label == 'Setpoint Corredera':
                    action[0] = value
                elif label == 'Setpoint Ángulo':
                    action[1] = value
                elif label == 'Setpoint Water':
                    action[2] = value
                elif label == 'Energía Motor Corredera':
                    action[3] = value
                elif label == 'Energía Motor Ángulo':
                    action[4] = value
                elif label == 'Energía Motor Válvula':
                    action[5] = value

                # Ejecutar un paso en el entorno con la acción modificada
                self.env.step(action)

                # Registrar el comando enviado
                command_str = f"Set {label} to {value} in environment\n"
                self.data_text_sent.insert(tk.END, command_str)
                self.data_text_sent.see(tk.END)
                print(command_str.strip())  # Para también imprimir en consola
        except ValueError as e:
            print(f"Invalid input: {e}")

    def send_pid_command(self, label_set):
        try:
            command = {
                'PID_Corredera': [0, 0, 0],
                'PID_Angulo': [0, 0, 0],
                'PID_Valvula': [0, 0, 0]
            }

            for label in label_set:
                value = float(self.entries[label].get())
                if 'Corredera' in label:
                    if 'Kp' in label:
                        command['PID_Corredera'][0] = value
                    elif 'Ki' in label:
                        command['PID_Corredera'][1] = value
                    elif 'Kd' in label:
                        command['PID_Corredera'][2] = value
                elif 'Ángulo' in label:
                    if 'Kp' in label:
                        command['PID_Angulo'][0] = value
                    elif 'Ki' in label:
                        command['PID_Angulo'][1] = value
                    elif 'Kd' in label:
                        command['PID_Angulo'][2] = value
                elif 'Válvula' in label:
                    if 'Kp' in label:
                        command['PID_Valvula'][0] = value
                    elif 'Ki' in label:
                        command['PID_Valvula'][1] = value
                    elif 'Kd' in label:
                        command['PID_Valvula'][2] = value

            if self.env:  # Si hay una instancia activa de BasicEnv
                self.env.set_pid_corredera(*command['PID_Corredera'])
                self.env.set_pid_angulo(*command['PID_Angulo'])
                self.env.set_pid_valvula(*command['PID_Valvula'])

            command_str = f"Setting PID values: {command}\n"
            self.data_text_sent.insert(tk.END, command_str)
            self.data_text_sent.see(tk.END)
        except ValueError as e:
            print(f"Invalid input: {e}")

    def toggle_manual_mode(self):
        manual_mode = self.manual_mode_var.get()
        if self.env:
            self.env.set_manual_mode(manual_mode)
        print(f"Setting manual mode to {'ON' if manual_mode else 'OFF'}")

    def calibrate_compass(self):
        if self.env:
            self.env.reset()  # Utilizar el método reset del entorno para calibrar
        print("Calibrating compass")

    def reset_robot(self):
        if self.env:
            self.env.reset()
        print("Resetting robot")

    ## FUNCIONES RELACIONADAS CON ENSAYOS LUEGO

    def setup_ensayos_tab(self):
        self.ensayos_treeview = ttk.Treeview(self.ensayos_frame)
        self.ensayos_treeview.pack(side='left', expand=True, fill='both')

        self.ensayos_treeview['columns'] = (
            "Nombre de la Planta", "Día", "Hora", "Tarea", "Regimen", "Magnitud", "Unidades", "Detalles")
        self.ensayos_treeview['show'] = 'headings'

        for col in self.ensayos_treeview['columns']:
            self.ensayos_treeview.heading(col, text=col)
            self.ensayos_treeview.column(col, width=120)

        self.actualizar_treeview_ensayos()

    def setup_plantas_section(self):
        self.left_frame = ttk.Frame(self.plantas_frame, width=200)
        self.left_frame.pack_propagate(0)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.cargar_eras(self.left_frame)

        self.right_frame = ttk.Frame(self.plantas_frame)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame = ttk.Frame(self.right_frame)
        self.buttons_frame.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.buttons_frame, text="Agregar Planta", command=self.agregar_planta).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.buttons_frame, text="Modificar Planta", command=self.modificar_planta).pack(side=tk.LEFT,
                                                                                                    padx=5)
        ttk.Button(self.buttons_frame, text="Eliminar Planta", command=self.eliminar_planta).pack(side=tk.LEFT, padx=5)

        self.plantas_treeview = ttk.Treeview(self.right_frame)
        self.plantas_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

    def setup_regimenes_tab(self):
        self.left_frame_regimenes = ttk.Frame(self.regimenes_frame, width=200)
        self.left_frame_regimenes.pack_propagate(0)
        self.left_frame_regimenes.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

        self.right_frame_regimenes = ttk.Frame(self.regimenes_frame)
        self.right_frame_regimenes.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.buttons_frame_regimenes = ttk.Frame(self.right_frame_regimenes)
        self.buttons_frame_regimenes.pack(side=tk.TOP, fill=tk.X)
        self.agregar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Agregar Tarea",
                                               command=self.agregar_tarea_regimen)
        self.agregar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.modificar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Modificar Tarea", state='disabled',
                                                 command=self.modificar_tarea_regimen)
        self.modificar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.eliminar_tarea_button = ttk.Button(self.buttons_frame_regimenes, text="Eliminar Tarea", state='disabled',
                                                command=self.eliminar_tarea_regimen)
        self.eliminar_tarea_button.pack(side=tk.LEFT, padx=5)

        self.regimenes_treeview = ttk.Treeview(self.right_frame_regimenes)
        self.regimenes_treeview.pack(side=tk.BOTTOM, expand=True, fill='both')

        self.regimenes_treeview['columns'] = ("Tarea", "Numero_Día", "Hora", "Magnitud", "Unidades", "Detalles")
        self.regimenes_treeview['show'] = 'headings'

        for col in self.regimenes_treeview['columns']:
            self.regimenes_treeview.heading(col, text=col)
            self.regimenes_treeview.column(col, width=120)

        self.regimenes_treeview.bind("<<TreeviewSelect>>", self.on_regimen_treeview_select)

        self.cargar_regimenes()


    def cargar_eras(self, container):
        for widget in container.winfo_children():
            widget.destroy()

        self.agregar_era_button = ttk.Button(container, text="Agregar Era", command=self.agregar_era, width=20)
        self.agregar_era_button.pack(side=tk.TOP, fill=tk.X)

        self.eliminar_era_button = ttk.Button(container, text="Eliminar Era", command=self.eliminar_era, width=20)
        self.eliminar_era_button.pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_plantas)

        for sheet_name in workbook.sheetnames:
            self.eras[sheet_name] = ttk.Button(container, text=sheet_name,
                                               command=lambda era=sheet_name: self.seleccionar_era(era))
            self.eras[sheet_name].pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def seleccionar_era(self, era):
        self.era_seleccionada = era
        self.mostrar_plantas_por_era(era)

    def mostrar_plantas_por_era(self, era):
        self.plantas_treeview.delete(*self.plantas_treeview.get_children())
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columns = ['Nombre de la Planta', 'Día Uno', 'Regimen', 'Velocidad de Agua', 'ÁnguloV', 'ÁnguloH',
                       'Posición X', 'Posición Y', 'Posición Z', 'Detalles']  # Asegúrate de incluir Detalles aquí
            self.plantas_treeview['columns'] = columns
            self.plantas_treeview['show'] = 'headings'
            for col in columns:
                self.plantas_treeview.heading(col, text=col)
                self.plantas_treeview.column(col, width=100)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.plantas_treeview.insert("", tk.END, values=row)
        else:
            messagebox.showerror("Error", f"La era '{era}' no tiene una hoja correspondiente en el libro de Excel.")

    def agregar_era(self):
        era_name = simpledialog.askstring("Nueva Era", "Nombre de la nueva Era:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name not in workbook.sheetnames:
                workbook.create_sheet(era_name)
                workbook.save(self.archivo_plantas)
                self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' ya existe.")

    def eliminar_era(self):
        era_name = simpledialog.askstring("Eliminar Era", "Nombre de la Era a eliminar:")
        if era_name:
            workbook = load_workbook(self.archivo_plantas)
            if era_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Confirmar", f"¿Estás seguro de que deseas eliminar la era '{era_name}'?")
                if confirm:
                    del workbook[era_name]
                    workbook.save(self.archivo_plantas)
                    self.cargar_eras(self.left_frame)
            else:
                messagebox.showerror("Error", f"La era '{era_name}' no existe.")

    def agregar_planta(self):
        add_window = tk.Toplevel(self)
        add_window.title("Agregar Nueva Planta")
        valores_actuales = {
            'Nombre de la Planta': '',
            'Regimen': '',
            'Dia Uno': '2024-01-01',
            'Posición X': 0,
            'Posición Y': 0,
            'Posición Z': 0,
            'Velocidad de Agua': 0,
            'Detalles': ''  # Asegúrate de incluir Detalles aquí
        }
        entries = {}
        for idx, (field, value) in enumerate(valores_actuales.items()):
            label = ttk.Label(add_window, text=field)
            label.grid(row=idx, column=0, padx=10, pady=10)
            entry_var = tk.StringVar(add_window, value=value)
            entry = ttk.Entry(add_window, textvariable=entry_var)
            entry.grid(row=idx, column=1, padx=10, pady=10)
            entries[field] = entry_var

        ttk.Button(add_window, text="Guardar Planta",
                   command=lambda: self.guardar_nueva_planta(entries, add_window)).grid(row=len(valores_actuales),
                                                                                        column=0, columnspan=2)

    def guardar_nueva_planta(self, entries, window):
        planta_details = {field: entry.get() for field, entry in entries.items()}

        if all(planta_details.values()):
            self.agregar_planta_a_excel(planta_details, self.era_seleccionada)
            window.destroy()
            self.mostrar_plantas_por_era(self.era_seleccionada)
        else:
            messagebox.showwarning("Agregar Planta", "Todos los campos son obligatorios.")

    def agregar_planta_a_excel(self, planta_details, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            sheet.append([
                planta_details['Nombre de la Planta'],
                planta_details['Regimen'],
                planta_details['Dia Uno'],
                planta_details['Posición X'],
                planta_details['Posición Y'],
                planta_details['Posición Z'],
                planta_details['Velocidad de Agua'],
                planta_details['Detalles']  # Asegúrate de incluir Detalles aquí
            ])
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def modificar_planta(self):
        item = self.plantas_treeview.selection()
        if item:
            selected_plant = self.plantas_treeview.item(item, 'values')
            modificar_window = tk.Toplevel(self)
            modificar_window.title("Modificar Planta")
            entries = {}
            for idx, (field, value) in enumerate(zip(self.plantas_treeview['columns'], selected_plant)):
                label = ttk.Label(modificar_window, text=field)
                label.grid(row=idx, column=0, padx=10, pady=10)
                entry_var = tk.StringVar(modificar_window, value=value)
                entry = ttk.Entry(modificar_window, textvariable=entry_var)
                entry.grid(row=idx, column=1, padx=10, pady=10)
                entries[field] = entry_var

            ttk.Button(modificar_window, text="Guardar Cambios",
                       command=lambda: self.guardar_cambios_planta(entries, modificar_window, item)).grid(
                row=len(selected_plant), column=0, columnspan=2)

    def guardar_cambios_planta(self, entries, window, item):
        new_values = [entry.get() for entry in entries.values()]
        self.plantas_treeview.item(item, values=new_values)

        self.actualizar_planta_en_excel(new_values, self.era_seleccionada)
        window.destroy()

    def actualizar_planta_en_excel(self, new_values, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                if row[0].value == new_values[0]:  # Asumiendo que el nombre de la planta es el primer campo
                    for idx, cell in enumerate(row):
                        cell.value = new_values[idx]
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def eliminar_planta(self):
        item = self.plantas_treeview.selection()
        if item:
            planta = self.plantas_treeview.item(item, 'values')
            confirm = messagebox.askyesno("Eliminar Planta", f"¿Deseas eliminar la planta '{planta[0]}'?")
            if confirm:
                self.eliminar_planta_de_excel(planta[0], self.era_seleccionada)
                self.plantas_treeview.delete(item)

    def eliminar_planta_de_excel(self, planta_name, era_seleccionada):
        workbook = load_workbook(self.archivo_plantas)
        if era_seleccionada in workbook.sheetnames:
            sheet = workbook[era_seleccionada]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                if row[0].value == planta_name:  # Asumiendo que el nombre de la planta es el primer campo
                    sheet.delete_rows(row[0].row, 1)
            workbook.save(self.archivo_plantas)
        else:
            messagebox.showerror("Error",
                                 f"La era '{era_seleccionada}' no tiene una hoja correspondiente en el libro de Excel.")

    def on_regimen_treeview_select(self, event):
        selected_item = self.regimenes_treeview.selection()
        if selected_item:
            self.modificar_tarea_button['state'] = 'normal'
            self.eliminar_tarea_button['state'] = 'normal'
        else:
            self.modificar_tarea_button['state'] = 'disabled'
            self.eliminar_tarea_button['state'] = 'disabled'

    def cargar_regimenes(self):
        for widget in self.left_frame_regimenes.winfo_children():
            widget.destroy()

        ttk.Button(self.left_frame_regimenes, text="Agregar Régimen",
                   command=self.agregar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        ttk.Button(self.left_frame_regimenes, text="Eliminar Régimen",
                   command=self.eliminar_regimen, width=20).pack(side=tk.TOP, fill=tk.X)

        workbook = load_workbook(self.archivo_regimenes)
        for regimen_name in workbook.sheetnames:
            ttk.Button(self.left_frame_regimenes, text=regimen_name,
                       command=lambda r=regimen_name: self.seleccionar_regimen(r),
                       width=20).pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    def agregar_regimen(self):
        regimen_name = simpledialog.askstring("Nuevo Régimen", "Nombre del nuevo Régimen:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(regimen_name)
                sheet.append(["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades",
                              "Detalles"])  # Asegúrate de incluir Detalles aquí
                workbook.save(self.archivo_regimenes)
                self.cargar_regimenes()
            else:
                messagebox.showerror("Error", f"El régimen '{regimen_name}' ya existe.")

    def eliminar_regimen(self):
        regimen_name = simpledialog.askstring("Eliminar Régimen", "Nombre del Régimen a eliminar:")
        if regimen_name:
            workbook = load_workbook(self.archivo_regimenes)
            if regimen_name in workbook.sheetnames:
                confirm = messagebox.askyesno("Eliminar", f"¿Estás seguro de eliminar el régimen '{regimen_name}'?")
                if confirm:
                    del workbook[regimen_name]
                    workbook.save(self.archivo_regimenes)
                    self.cargar_regimenes()
            else:
                messagebox.showerror("Error", "Régimen no encontrado.")

    def seleccionar_regimen(self, regimen):
        self.regimen_actual = regimen
        self.mostrar_tareas_por_regimen(regimen)

    def agregar_tarea_regimen(self):
        regimen = self.regimen_actual

        ventana_emergente = tk.Toplevel(self)
        ventana_emergente.title("Agregar Nueva Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "Tiempo de ejecución (s)", "Magnitud", "Unidades",
                  "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}

        for idx, campo in enumerate(campos):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, sticky="w")
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if regimen in workbook.sheetnames:
                sheet = workbook[regimen]
                sheet.append(list(detalles_tarea.values()))
                workbook.save(self.archivo_regimenes)
                ventana_emergente.destroy()
                self.mostrar_tareas_por_regimen(regimen)
            else:
                messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Agregar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                    columnspan=2, pady=10)

    def modificar_tarea_regimen(self):
        selected_item = self.regimenes_treeview.selection()[0]
        tarea_seleccionada = self.regimenes_treeview.item(selected_item, "values")

        ventana_emergente = tk.Toplevel(self)
        ventana_emergente.title("Modificar Tarea")

        campos = ["Tarea", "Numero_Día", "Hora", "tiempo_ejecución(s)", "magnitud", "unidades",
                  "Detalles"]  # Asegúrate de incluir Detalles aquí
        entradas = {}
        for idx, (campo, valor) in enumerate(zip(campos, tarea_seleccionada)):
            ttk.Label(ventana_emergente, text=campo).grid(row=idx, column=0, padx=10, pady=5)
            entrada = ttk.Entry(ventana_emergente)
            entrada.grid(row=idx, column=1, padx=10, pady=5)
            entrada.insert(0, valor)
            entradas[campo] = entrada

        def confirmar():
            detalles_tarea_modificada = {campo: entradas[campo].get() for campo in campos}

            workbook = load_workbook(self.archivo_regimenes)
            if self.regimen_actual in workbook.sheetnames:
                sheet = workbook[self.regimen_actual]
                fila_tarea = self.regimenes_treeview.index(selected_item) + 2
                for idx, valor in enumerate(detalles_tarea_modificada.values(), start=1):
                    sheet.cell(row=fila_tarea, column=idx, value=valor)
                workbook.save(self.archivo_regimenes)
                self.regimenes_treeview.item(selected_item, values=list(detalles_tarea_modificada.values()))
                ventana_emergente.destroy()
            else:
                messagebox.showerror("Error", f"El régimen '{self.regimen_actual}' no existe.")
                ventana_emergente.destroy()

        ttk.Button(ventana_emergente, text="Modificar Tarea", command=confirmar).grid(row=len(campos), column=0,
                                                                                      columnspan=2, pady=10)

    def eliminar_tarea_regimen(self, regimen):
        selected = self.regimenes_treeview.selection()
        if selected:
            tarea_id = self.regimenes_treeview.index(selected[0]) + 2
            workbook = load_workbook(self.archivo_regimenes)
            sheet = workbook[regimen]
            sheet.delete_rows(tarea_id)
            workbook.save(self.archivo_regimenes)
            self.mostrar_tareas_por_regimen(regimen)

    def mostrar_tareas_por_regimen(self, regimen):
        try:
            if self.regimenes_treeview.winfo_exists():
                self.regimenes_treeview.delete(*self.regimenes_treeview.get_children())
                workbook = load_workbook(self.archivo_regimenes)
                if regimen in workbook.sheetnames:
                    sheet = workbook[regimen]
                    if not self.regimenes_treeview['columns']:
                        self.regimenes_treeview['columns'] = [sheet.cell(row=1, column=j).value for j in
                                                              range(1, sheet.max_column + 1)]
                        for col in self.regimenes_treeview['columns']:
                            self.regimenes_treeview.heading(col, text=col)
                            self.regimenes_treeview.column(col, width=100)
                    for i in range(2, sheet.max_row + 1):
                        row_values = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
                        self.regimenes_treeview.insert("", tk.END, values=row_values)
                else:
                    messagebox.showerror("Error", f"El régimen '{regimen}' no existe.")
            self.configurar_botones_accion_regimen(regimen)

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al mostrar las tareas: {e}")

    def configurar_botones_accion_regimen(self, regimen):
        if not hasattr(self, 'agregar_tarea_button'):
            self.agregar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Agregar Tarea a " + regimen,
                                                   command=lambda: self.agregar_tarea_regimen())
            self.agregar_tarea_button.pack(padx=10, pady=5)

            self.eliminar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Eliminar Tarea de " + regimen,
                                                    command=lambda: self.eliminar_tarea_regimen(regimen),
                                                    state='disabled')
            self.eliminar_tarea_button.pack(padx=10, pady=5)

            self.modificar_tarea_button = ttk.Button(self.right_frame_regimenes, text="Modificar Tarea de " + regimen,
                                                     command=lambda: self.modificar_tarea_regimen(),
                                                     state='disabled')
            self.modificar_tarea_button.pack(padx=10, pady=5)
        else:
            self.agregar_tarea_button.config(text="Agregar Tarea a " + regimen,
                                             command=lambda: self.agregar_tarea_regimen())
            self.eliminar_tarea_button.config(text="Eliminar Tarea de " + regimen,
                                              command=lambda: self.eliminar_tarea_regimen(regimen))
            self.modificar_tarea_button.config(text="Modificar Tarea de " + regimen,
                                               command=lambda: self.modificar_tarea_regimen())

    def on_regimen_treeview_select(self, event):
        selected = self.regimenes_treeview.selection()
        if selected:
            self.eliminar_tarea_button['state'] = 'normal'
            self.modificar_tarea_button['state'] = 'normal'
        else:
            self.eliminar_tarea_button['state'] = 'disabled'
            self.modificar_tarea_button['state'] = 'disabled'

    def actualizar_treeview_ensayos(self):
        tareas_df = pd.read_excel(self.archivo_ensayos)

        for i in self.ensayos_treeview.get_children():
            self.ensayos_treeview.delete(i)

        for _, fila in tareas_df.iterrows():
            self.ensayos_treeview.insert(
                '', 'end',
                values=(
                    fila['Nombre de la Planta'],
                    fila['Día'],
                    fila['Hora'],
                    fila['Tarea'],
                    fila['Regimen'],
                    fila['Magnitud'],
                    fila['Unidades'],
                    fila['Detalles']  # Asegurarse de incluir Detalles aquí
                )
            )

if __name__ == "__main__":
    env = BasicEnv()  # Crea la instancia del entorno BasicEnv
    app = Interfaz('archivo_plantas.xlsx', 'archivo_regimenes.xlsx', 'archivo_ensayos.xlsx', env=env)
    app.mainloop()
