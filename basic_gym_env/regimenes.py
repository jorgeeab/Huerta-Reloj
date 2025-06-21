from openpyxl import load_workbook, Workbook

class RegimenesManager:
    def __init__(self, archivo_regimenes):
        self.archivo_regimenes = archivo_regimenes
        self.verificar_o_crear_archivo_regimenes()

    def verificar_o_crear_archivo_regimenes(self):
        # Nombres de las columnas que deberían estar presentes
        columnas_correctas = ["Tarea", "Numero_Día", "Hora", "Tiempo de Ejecución (s)", "Magnitud", "Unidades"]
        try:
            workbook = load_workbook(self.archivo_regimenes)
            # Verificar si hay hojas en el archivo
            if not workbook.sheetnames:
                raise ValueError("El archivo no tiene hojas")
            # Verificar si las columnas están en la primera hoja
            first_sheet = workbook[workbook.sheetnames[0]]
            columnas_existentes = [cell.value for cell in first_sheet[1]]
            if columnas_existentes != columnas_correctas:
                raise ValueError("Columnas incorrectas")
        except (FileNotFoundError, ValueError, KeyError, Exception):
            # Crear un nuevo archivo si no existe o está dañado
            print("Archivo de regímenes no encontrado o incorrecto. Creando uno nuevo...")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Regimen 1"  # Nombre del primer régimen por defecto
            sheet.append(columnas_correctas)
            workbook.save(self.archivo_regimenes)

    # Cargar los regímenes existentes
    def cargar_regimenes(self):
        workbook = load_workbook(self.archivo_regimenes)
        return workbook.sheetnames

    # Listar las tareas de un régimen específico
    def listar_tareas_de_regimen(self, regimen):
        workbook = load_workbook(self.archivo_regimenes)
        tareas = []
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorar el encabezado
                tareas.append({
                    'Tarea': row[0],
                    'Numero_Día': row[1],
                    'Hora': row[2],
                    'Tiempo de Ejecución (s)': row[3],
                    'Magnitud': row[4],
                    'Unidades': row[5]
                })
        return tareas

    # **Nuevo método: Obtener detalles de un régimen específico**
    def obtener_regimen(self, regimen):
        # Retorna todas las tareas del régimen
        return self.listar_tareas_de_regimen(regimen)

    def agregar_regimen(self, regimen_name):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(regimen_name)
            sheet.append(["Tarea", "Numero_Día", "Hora", "Tiempo de Ejecución (s)", "Magnitud", "Unidades"])
            workbook.save(self.archivo_regimenes)

    def agregar_tarea(self, regimen, tarea_details):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            sheet.append([
                tarea_details['Tarea'],
                tarea_details['Numero_Día'],
                tarea_details['Hora'],
                tarea_details['Tiempo de Ejecución (s)'],
                tarea_details['Magnitud'],
                tarea_details['Unidades']
            ])
            workbook.save(self.archivo_regimenes)

    def modificar_tarea(self, regimen, fila, updated_values):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            columnas = [cell.value for cell in sheet[1]]  # Obtener nombres de las columnas
            for i, key in enumerate(columnas, start=1):
                if key in updated_values:
                    sheet.cell(row=fila, column=i).value = updated_values[key]
            workbook.save(self.archivo_regimenes)

    def eliminar_tarea(self, regimen, fila):
        workbook = load_workbook(self.archivo_regimenes)
        if regimen in workbook.sheetnames:
            sheet = workbook[regimen]
            sheet.delete_rows(fila)
            workbook.save(self.archivo_regimenes)
