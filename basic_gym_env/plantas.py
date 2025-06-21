from openpyxl import load_workbook, Workbook

class PlantasManager:
    def __init__(self, archivo_plantas):
        self.archivo_plantas = archivo_plantas
        self.verificar_o_crear_archivo_plantas()

    def verificar_o_crear_archivo_plantas(self):
        # Nombres de las columnas que deberían estar presentes
        columnas_correctas = [
            'Nombre de la Planta', 'Regimen', 'Dia Uno', 'Posición X',
            'Posición Y', 'Posición Z', 'Velocidad de Agua'
        ]
        try:
            workbook = load_workbook(self.archivo_plantas)
            # Verificar si hay hojas en el archivo
            if not workbook.sheetnames:
                raise ValueError("El archivo no tiene hojas")
            # Verificar si las columnas están en la primera hoja
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                columnas_existentes = [cell.value for cell in sheet[1]]
                if columnas_existentes != columnas_correctas:
                    raise ValueError("Columnas incorrectas en la hoja " + sheet_name)
        except (FileNotFoundError, ValueError, KeyError, Exception):
            # Crear un nuevo archivo si no existe o está dañado
            print("Archivo de plantas no encontrado o incorrecto. Creando uno nuevo...")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Era 1"  # Nombre de la primera era por defecto
            sheet.append(columnas_correctas)
            workbook.save(self.archivo_plantas)

    # Cargar las eras existentes
    def cargar_eras(self):
        workbook = load_workbook(self.archivo_plantas)
        return workbook.sheetnames

    # Listar todas las plantas de una era específica
    def listar_plantas_de_era(self, era):
        workbook = load_workbook(self.archivo_plantas)
        plantas = []
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columnas = [cell.value for cell in sheet[1]]  # Nombres de las columnas
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorar el encabezado
                planta = dict(zip(columnas, row))
                plantas.append(planta)
        return plantas

    # Obtener el régimen de una planta en una era específica
    def obtener_regimen_de_planta(self, era, nombre_planta):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == nombre_planta:
                    return row[1]  # Regimen está en la columna 2
        return None

    # Obtener detalles de una planta específica
    def obtener_planta(self, era, nombre_planta):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columnas = [cell.value for cell in sheet[1]]  # Obtener nombres de las columnas
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == nombre_planta:
                    planta = dict(zip(columnas, row))
                    return planta  # Retorna un diccionario con los detalles de la planta
        return None

    def agregar_planta(self, planta_details, era):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
        else:
            # Crear una nueva era si no existe
            sheet = workbook.create_sheet(era)
            sheet.append([
                'Nombre de la Planta', 'Regimen', 'Dia Uno', 'Posición X',
                'Posición Y', 'Posición Z', 'Velocidad de Agua'
            ])
        sheet.append([
            planta_details['Nombre de la Planta'],
            planta_details['Regimen'],
            planta_details['Dia Uno'],
            planta_details['Posición X'],
            planta_details['Posición Y'],
            planta_details['Posición Z'],
            planta_details['Velocidad de Agua']
        ])
        workbook.save(self.archivo_plantas)

    def modificar_planta(self, era, fila, updated_values):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            columnas = [cell.value for cell in sheet[1]]  # Obtener nombres de las columnas
            for i, key in enumerate(columnas, start=1):
                if key in updated_values:
                    sheet.cell(row=fila, column=i).value = updated_values[key]
            workbook.save(self.archivo_plantas)

    def eliminar_planta(self, era, fila):
        workbook = load_workbook(self.archivo_plantas)
        if era in workbook.sheetnames:
            sheet = workbook[era]
            sheet.delete_rows(fila)
            workbook.save(self.archivo_plantas)

    # Listar todas las plantas en todas las eras
    def listar_todas_las_plantas(self):
        workbook = load_workbook(self.archivo_plantas)
        all_plantas = []
        for era in workbook.sheetnames:
            sheet = workbook[era]
            columnas = [cell.value for cell in sheet[1]]  # Obtener nombres de las columnas
            for row in sheet.iter_rows(min_row=2, values_only=True):
                planta = dict(zip(columnas, row))
                planta['Era'] = era  # Agregar la era al diccionario
                all_plantas.append(planta)
        return all_plantas
  # Método para obtener todas las actividades de cuidado
    def obtener_todas_las_actividades(self):
        workbook = load_workbook(self.archivo_plantas)
        actividades = []
        for era in workbook.sheetnames:
            sheet = workbook[era]
            columnas = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                planta = dict(zip(columnas, row))
                regimen = planta.get('Regimen')
                nombre_planta = planta.get('Nombre de la Planta')
                # Supongamos que 'Regimen' contiene las actividades de cuidado en un formato estructurado
                # Por ejemplo, una lista de tuplas: [(fecha, actividad), ...]
                # Necesitas parsear esta información según corresponda
                actividades_planta = self.parsear_regimen(regimen, nombre_planta)
                actividades.extend(actividades_planta)
        return actividades

    def parsear_regimen(self, regimen, nombre_planta):
        # Este método debe parsear el régimen y devolver una lista de actividades
        # Por ejemplo, si el régimen es una cadena como "Regar cada 3 días a partir del 2023-01-01"
        # Debes parsearlo y generar las fechas correspondientes
        actividades = []
        # Implementa la lógica de parseo aquí
        # Para demostración, supongamos que el régimen es una lista de diccionarios
        # e.g., [{'fecha': '2023-01-01', 'actividad': 'Regar'}, ...]
        # Puedes modificar esto según tu formato de datos real
        for actividad in regimen:
            fecha = actividad.get('fecha')
            accion = actividad.get('actividad')
            actividades.append({
                'Fecha': fecha,
                'Actividad': accion,
                'Planta': nombre_planta
            })
        return actividades

    def generar_calendario(self):
        actividades = self.obtener_todas_las_actividades()
        # Ordenar las actividades por fecha
        actividades.sort(key=lambda x: x['Fecha'])
        # Exportar a una nueva hoja de Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Calendario de Cuidados'
        sheet.append(['Fecha', 'Planta', 'Actividad'])
        for actividad in actividades:
            sheet.append([actividad['Fecha'], actividad['Planta'], actividad['Actividad']])
        workbook.save('Calendario_Cuidados.xlsx')